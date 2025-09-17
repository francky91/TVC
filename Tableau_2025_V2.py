import tkinter as tk
from tkinter import ttk, messagebox
import json
import math
import sys
from utils import *
from json_access import load_all_tableau, save_tableau
import os
import pandas as pd
from openpyxl import load_workbook
import PIL
from SaisieMatch import validate_input_set,handle_mode_change,update_set_visibility

class TournamentBracket:
    def __init__(self, root, category, round_number):

        self.root = root
        self.category = category
        self.round_number = round_number
        self.tourInitial = 0
        self.nb_tops = 0
        self.nb_poules = 0
        self.fichier_Tableau = ""
        self.results_modified = [False]
        self.previous_mode = "2 sets"
        self.checkbox_vars = {}
        self.ok_window_open = False  # État pour la fenêtre "Tournoi OK"
        self.ko_window_open = False  # État pour la fenêtre "Tournoi KO"
        self.intertops_window = None  # Référence à la fenêtre "InterTops"
        self.intertops_window_open = False  # État pour la fenêtre "InterTops"
        
        self.root.title(f"Tournoi {category} Tour {round_number}")

        '''
        # Conteneur pour le canevas
        self.canvas_frame = tk.Frame(root)
        self.canvas_frame.pack(fill="both", expand=True)

        # Canvas pour dessiner le tableau avec défilement
        self.canvas = tk.Canvas(self.canvas_frame, width=1000, height=600, bg="white")
        self.scrollbar = tk.Scrollbar(self.canvas_frame, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.scrollbar.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)

        frame_buttons = ttk.Frame(self.root)
        frame_buttons.pack(fill="x", padx=10, pady=5)

        # Bouton pour "Charger le Tournoi OK"
        self.load_button_ok = ttk.Button(frame_buttons, text="Charger le Tournoi OK", command=lambda: self.load_tournament("Tournoi OK"))
        self.load_button_ok.pack(side="left", padx=5)

        # Bouton pour "Charger le Tournoi KO"
        self.load_button_ko = ttk.Button(frame_buttons, text="Charger le Tournoi KO", command=lambda: self.load_tournament("Tournoi KO"))
        self.load_button_ko.pack(side="left", padx=5)

        # Bouton pour "InterTops"
        self.inter_tops_button = ttk.Button(frame_buttons, text="InterTops", command=lambda: self.load_tournament("InterTops"))
        self.inter_tops_button.pack(side="left", padx=5)
        
        self.calcul_points_button = ttk.Button(frame_buttons, text="Calcul de points", command=self.show_points_window)
        self.calcul_points_button.pack(side="left", padx=5)
        '''
        
        # Conteneur pour les boutons
        frame_buttons = ttk.Frame(self.root)
        frame_buttons.pack(fill="both", expand=True, padx=20, pady=20)

        # Bouton pour "Charger le Tournoi OK"
        self.load_button_ok = ttk.Button(frame_buttons, text="Charger le Tournoi OK", command=lambda: self.load_tournament("Tournoi OK"))
        self.load_button_ok.pack(fill="x", pady=5)

        # Bouton pour "Charger le Tournoi KO"
        self.load_button_ko = ttk.Button(frame_buttons, text="Charger le Tournoi KO", command=lambda: self.load_tournament("Tournoi KO"))
        self.load_button_ko.pack(fill="x", pady=5)

        # Bouton pour "InterTops"
        self.inter_tops_button = ttk.Button(frame_buttons, text="InterTops", command=lambda: self.load_tournament("InterTops"))
        self.inter_tops_button.pack(fill="x", pady=5)

        # Bouton pour "Calcul de points"
        self.calcul_points_button = ttk.Button(frame_buttons, text="Calcul de points", command=self.show_points_window)
        self.calcul_points_button.pack(fill="x", pady=5)

        # Ajuster la taille minimale de la fenêtre principale
        self.root.geometry("300x200")  # Largeur x Hauteur
        
        self.tableauOK = []
        self.tableauKO = []
        
        self.data = {}
        
        self.read_config_files()
        
        self.debugStatus = lire_status_debug("config.ini")
                

    def extract_players_by_progression(self, tableau, nbPoints, players_points):
        """
        Extrait les joueurs d'un tableau en suivant la progression des rounds.
        Associe directement chaque joueur au nombre de points dans players_points.
        Retourne la nouvelle valeur de nbPoints une fois la boucle terminée.
        """
        level = 0
        fini = False
        firstRoundFound = False

        while not fini:
            string_level = f"1/{level}"
            #print(f"level: {level} string_level: {string_level}")
            # Si on a déjà trouvé un round et que le suivant n'existe pas, on arrête.
            if firstRoundFound and string_level not in tableau:
                fini = True
            else:
                if string_level in tableau:
                    firstRoundFound = True
                    for key, joueurs in tableau[string_level].items():
                        # Joueur1
                        j1 = joueurs.get('joueur1', "")
                        if j1 and j1 != "---" and j1 not in players_points:
                            players_points[j1] = nbPoints
                            print(f"joueur1: {j1} points: {nbPoints}, level: {level}")
                            
                        # Joueur2
                        j2 = joueurs.get('joueur2', "")
                        if j2 and j2 != "---" and j2 not in players_points:
                            players_points[j2] = nbPoints
                            print(f"joueur2: {j2} points: {nbPoints}, level: {level}")

                # Progression de level et mise à jour de nbPoints
                if level == 0:
                    level = 1
                    nbPoints -= 2
                else:
                    level *= 2                                      
                    nbPoints -= 1
        
        return (nbPoints + 2)

    def show_points_window(self):
        """Affiche une fenêtre pour gérer les points des joueurs."""

        def validate_int_0_15(new_text):
            """
            Retourne True si new_text est vide (on autorise la suppression)
            ou si c'est un entier entre 0 et 15. Sinon False.
            """
            # Autoriser une chaîne vide (permet de supprimer pour retaper)
            if new_text == "":
                return True

            # Vérifier que c'est bien un nombre entier
            if new_text.isdigit():
                val = int(new_text)
                return 0 <= val <= 15
        
            return False        
        
        # Dictionnaire qui contiendra { "NomDuJoueur": points_calcules, ... }
        players_points = {}

        # Déterminer le nombre de points max initial
        if self.nb_poules <= 8:
            nbPointsMax = 10
        elif self.nb_poules <= 16:
            nbPointsMax = 12
        elif self.nb_poules <= 32:
            nbPointsMax = 14

        # Si on a des Tops, on ajoute 1 point supplémentaire
        if self.nb_tops > 0:
            nbPointsMax += 1

        # Récupérer joueurs+points depuis tableauOK et tableauKO
        nbPointsMax = self.extract_players_by_progression(self.tableauOK, nbPointsMax, players_points)
        nbPointsMax = self.extract_players_by_progression(self.tableauKO, nbPointsMax, players_points)

        self.points_entries = {}

        # Fenêtre toplevel
        points_window = tk.Toplevel(self.root)
        points_window.title("Calcul des Points")

        # Fixer une taille minimale ou initiale pour bien voir les boutons
        points_window.geometry("600x400")

        # -- Configuration pour faire deux lignes (row=0 -> zone défilable, row=1 -> boutons) --
        points_window.rowconfigure(0, weight=1)  
        points_window.columnconfigure(0, weight=1)

        # -- 1) Zone défilable en row=0 --
        scrollable_container = ttk.Frame(points_window)
        scrollable_container.grid(row=0, column=0, sticky="nsew")

        canvas = tk.Canvas(scrollable_container)
        canvas.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(scrollable_container, orient="vertical", command=canvas.yview)
        scrollbar.pack(side="right", fill="y")

        canvas.configure(yscrollcommand=scrollbar.set)

        scrollable_frame = ttk.Frame(canvas)
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")    

        validate_cmd = points_window.register(validate_int_0_15)        
        
        # Ajout des joueurs et champs de saisie (avec la valeur calculée)
        for row, (joueur, points_calc) in enumerate(players_points.items(), start=1):
            ttk.Label(scrollable_frame, text=joueur, anchor="w").grid(row=row, column=0, padx=5, pady=5, sticky="w")

            # Variable de saisie initialisée avec le nb de points
            points_var = tk.StringVar(value=str(points_calc))
            points_entry = ttk.Entry(
                scrollable_frame, 
                textvariable=points_var, 
                width=10,
                validate="key",                   # validation sur chaque frappe
                validatecommand=(validate_cmd, "%P")  # "%P" = nouveau texte proposé
                )
            points_entry.grid(row=row, column=1, padx=5, pady=5)

            # On garde la référence pour pouvoir relire ces points plus tard
            self.points_entries[joueur] = points_var

        buttons_frame = ttk.Frame(points_window)
        buttons_frame.grid(row=1, column=0, sticky="ew", padx=5, pady=5)

        validate_button = ttk.Button(buttons_frame, text="Valider", command=self.save_points)
        validate_button.pack(side="left", padx=5)
        
        quit_button = ttk.Button(buttons_frame, text="Fermer", command=points_window.destroy)
        quit_button.pack(side="right", padx=5)     
        
    def quit_match(self, window):
        if (self.results_modified[0]):
            response = messagebox.askyesno(
                    "Info",
                    f"Ders données n'ont pas été sauvées, continuez-vous ?")
            if response:
                self.results_modified = [False]
                window.destroy()
            else:
                window.focus_set()
        else:
            window.destroy()
        
    def save_points(self):
        """
        Exemple de méthode pour récupérer et afficher les points.
        Vous pouvez ensuite les stocker en base ou dans un fichier.
        """
        config_file = "config.ini"
        try:
            fichier_entree = lire_excel_name()
            print(f"Fichier d'entrée : {fichier_entree}")
            
            df = pd.read_excel(fichier_entree, sheet_name=self.category, header=6)
            
            wb = load_workbook(fichier_entree)
            feuille = wb[self.category]
            tour = int(self.round_number)
            champ_presence = f"Pr. T{tour}"
            df_filtre = df[df[champ_presence].str.lower() == "x"]
           
            colonnes_initiale_presence = 8
            colonne_initiale_points =9
            ligne = 8
            
            
            colonne_presence = colonnes_initiale_presence + (tour - 1)*2
            print (f"colonne_presence:{colonne_presence} tour:{tour} colonne_initiale_points:{colonne_initiale_points}")
            colonne_point = colonne_initiale_points + (tour - 1)*2
            print(f"colonne_point:{colonne_point} tour:{tour} colonne_initiale_points:{colonne_initiale_points}")
            
            colonne_nom = 2
            colonne_prenom = 3
            colonne_club = 6
            print(f"self.points_entries:{self.points_entries}")
            for row in range(2, feuille.max_row + 1):  # on part de la ligne 2 car la 1 est l'entête
                if feuille.cell(row=row, column=colonne_presence).value == "X" or  feuille.cell(row=row, column=colonne_presence).value == "x":
                    nom_val = feuille.cell(row=row, column=colonne_nom).value
                    prenom_val = feuille.cell(row=row, column=colonne_prenom).value
                    club_val = feuille.cell(row=row, column=colonne_club).value
                    print(f"nom_val:{nom_val} prenom_val:{prenom_val} club_val:{club_val}")
                    cle = f"{nom_val} {prenom_val} {club_val}"
                
                    champ_points = f"Points T{tour}"

                    if (cle in self.points_entries):
                        
                        points = (int)(self.points_entries[cle].get())
                        print(f"points:{points} cle:{cle}")
                        if (points != None):
                            points_feuille = feuille.cell(row=row, column=colonne_point).value
                            print(f"points_feuille:{points_feuille} points:{points}")
                            if (points_feuille != None):
                                points_feuille = int(points_feuille)
                            if (points_feuille == None):
                                print("Ecrit")
                                feuille.cell(row=row, column=colonne_point).value = points
                            elif (points != points_feuille ):
                                print("on est ici")
                                response = messagebox.askyesno(        
                                "Info",
                                f"Il existe déjà des valeurs pour {cle}, continuez-vous ?"
                                )
                                print(f"response:{response}")
                                if (response):
                                    feuille.cell(row=row, column=colonne_point).value = points                        
                            
            wb.save(fichier_entree)
            messagebox.showinfo(title="Info", message="Points joueurs sauvegardés")
            self.results_modified = [True]
            #window.destroy()
        except FileNotFoundError as e:
            print(e)
        except Exception as e:
            print(f"Erreur : {e}")
            
    def autosave_data(self):
        """
        Sauvegarde silencieuse : on enregistre dans le JSON, 
        mais on ne ferme aucune fenêtre et on ne recharge pas le tableau.
        """
        data = {
            "category": self.category,
            "round": self.round_number,
            "tourInitial": self.tourInitial,
            "nbTops": self.nb_tops,
            "nbPoules": self.nb_poules,
            "Tableau OK": self.tableauOK,
            "Tableau KO": self.tableauKO
        }
        resp = save_tableau(self.fichier_Tableau, data)
        if resp:
            print("Sauvegarde automatique OK.")
            self.results_modified[0] = False
        else:
            print("Erreur lors de la sauvegarde automatique.")
    
    def load_tournament(self, Tableau):
        # Vérifier si la fenêtre correspondante est déjà ouverte
        if Tableau == "Tournoi OK" and self.ok_window_open:
            self.ok_window.focus_set()  # Donner le focus à la fenêtre "Tournoi OK"
            return
        elif Tableau == "Tournoi KO" and self.ko_window_open:
            self.ko_window.focus_set()  # Donner le focus à la fenêtre "Tournoi KO"
            return
        elif Tableau == "InterTops" and hasattr(self, "intertops_window_open") and self.intertops_window_open:
            self.intertops_window.focus_set()  # Donner le focus à la fenêtre "InterTops"
            return
        
        if self.results_modified[0]:
            self.autosave_data()
        
        self.fichier_Tableau = f"Tableau_{self.category}_tour{self.round_number}.json"
        self.data = load_all_tableau(self.fichier_Tableau, True)
        self.tourInitial = int(self.data.get("tourInitial"))
        self.nb_tops = int(self.data.get("nbTops"))
        self.nb_poules = int(self.data.get("nbPoules"))
        
        if self.nb_tops < 2:
            self.inter_tops_button.config(state="disabled")   
        
        self.tableauOK = self.data["Tableau OK"]
        self.tableauKO = self.data["Tableau KO"]
        
        self.current_table = Tableau
        
        # Créer une nouvelle fenêtre Toplevel
        new_window = tk.Toplevel(self.root)
        new_window.title(f"{Tableau} - {self.category} Tour {self.round_number}")
        new_window.geometry("1200x800")  # Taille de la fenêtre

        # Ajouter un label pour indiquer le type de tableau
        label = ttk.Label(new_window, text=f"{self.category} - {Tableau}", font=("Arial", 16, "bold"))
        label.pack(pady=10)

        # Conteneur pour le canevas
        canvas_frame = tk.Frame(new_window)
        canvas_frame.pack(fill="both", expand=True)

        # Canvas pour dessiner le tableau avec défilement
        canvas = tk.Canvas(canvas_frame, width=1000, height=600, bg="white")
        scrollbar = tk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        
        # Ajouter le bouton "PDF" dans la fenêtre
        if (Tableau == "Tournoi OK" or Tableau == "Tournoi KO"):
            pdf_button = ttk.Button(new_window, text="Exporter en PDF", command=lambda: self.print_bracket_to_pdf(canvas))
            pdf_button.pack(pady=10)
				
        firstTour=self.tourInitial
        if (self.nb_tops >= 1):
            firstTour= firstTour * 2
            
        if Tableau == "Tournoi OK":
            # Titre : "Tournoi OK {catégorie} Tour {round_number}"
            # new_title = f"Tournoi OK {self.category} Tour {self.round_number}"
            # self.root.title(new_title)
            #self.draw_bracket(self.tableauOK, firstTour, self.nb_tops * 4, table_name="OK")
            self.draw_bracket(self.tableauOK, firstTour, self.nb_tops * 4, table_name="OK", canvas=canvas)
            self.ok_window_open = True  # Marquer la fenêtre comme ouverte
            self.ok_window = new_window  # Stocker la référence à la fenêtre
            self.ok_canvas = canvas  # Stocker le canevas pour le PDF

        elif Tableau == "Tournoi KO":
            # Titre : "Tournoi KO {catégorie} Tour {round_number}"
            # new_title = f"Tournoi KO {self.category} Tour {self.round_number}"
            # self.root.title(new_title)
            # self.draw_bracket(self.tableauKO, self.tourInitial, 0, table_name="KO")
            self.draw_bracket(self.tableauKO, self.tourInitial, 0, table_name="KO", canvas=canvas)
            self.ko_window_open = True  # Marquer la fenêtre comme ouverte
            self.ko_window = new_window  # Stocker la référence à la fenêtre
            self.ko_canvas = canvas  # Stocker le canevas pour le PDF

        elif Tableau == "InterTops":
            # Titre : "Intertops {catégorie} Tour {round_number}"
            #new_title = f"Intertops {self.category} Tour {self.round_number}"
            #self.root.title(new_title)
            self.draw_intertops(self.tableauOK, canvas=canvas)
            self.intertops_window_open = True  # Marquer la fenêtre comme ouverte
            self.intertops_window = new_window  # Stocker la référence à la fenêtre
            self.intertops_canvas = canvas  # Stocker le canevas pour le PDF
            
                # Réinitialiser l'état lorsque la fenêtre est fermée
        def on_close():
            if Tableau == "Tournoi OK":
                self.ok_window_open = False
                self.ok_window = None  # Réinitialiser la référence
                self.ok_canvas = None  # Réinitialiser le canevas
            elif Tableau == "Tournoi KO":
                self.ko_window_open = False
                self.ko_window = None  # Réinitialiser la référence
                self.ko_canvas = None  # Réinitialiser le canevas
            elif Tableau == "InterTops":
                self.intertops_window_open = False
                self.intertops_window = None  # Réinitialiser la référence
                self.intertops_canvas = None  # Réinitialiser le canevas
            new_window.destroy()

        new_window.protocol("WM_DELETE_WINDOW", on_close)
    
    def set_new_player(self, level, idx, player, nb_tops, dossard, tableau, intertop=False):
            new_level = ""
            try:
                if (not intertop):
                    current_level = int(level.split("/")[1])  # Extrait la partie numérique
                    next_level = current_level // 2
                    new_level = f"1/{next_level}"
                else:
                    new_level = "1/8"
            except (ValueError, IndexError):
                print(f"Erreur : Niveau '{level}' est invalide.")
            # Initialiser le niveau dans tableauOK s'il n'existe pas
            if new_level not in tableau:
                tableau[new_level] = {}
            
            if (not intertop):
                if (int(current_level) == 2*nb_tops):
                    index_table = str(idx)
                else:
                    index_table = str(1 + ((idx - 1) // 2))
                # Initialiser l'entrée pour ce match si elle n'existe pas
                if str(index_table) not in tableau[new_level]:
                    tableau[new_level][index_table] = {}

                # Mettre à jour ou ajouter le joueur dans le niveau suivant
                if (int(current_level) == 2*nb_tops):
                    print(f"tableau[{new_level}][{index_table}][joueur2]={player}")
                    tableau[new_level][str(index_table)]["joueur2"] = player
                    tableau[new_level][str(index_table)]["dossard_joueur2"] = dossard
                else:
                    if (idx %2 == 1):
                        tableau[new_level][str(index_table)]["joueur1"] = player
                        print(f"tableau[{new_level}][{index_table}][joueur1]={player}")
                        tableau[new_level][str(index_table)]["dossard_joueur1"] = dossard
                    else:
                        tableau[new_level][str(index_table)]["joueur2"] = player
                        print(f"tableau[{new_level}][{index_table}][joueur2]={player}")
                        tableau[new_level][str(index_table)]["dossard_joueur2"] = dossard
            else:
                if str(idx) not in tableau[new_level]:
                    tableau[new_level][str(idx)] = {}
                tableau[new_level][str(idx)]["joueur1"] = player
                tableau[new_level][str(idx)]["dossard_joueur1"] = dossard

    def gestion_vainqueur(self, player1, player2, dossard1, dossard2, level, index_match, tableau, nb_tops, intertop, scores_sets, match_window):
        if (intertop):
            #table_intertops = [(6,3), (7, 2), (5, 4), (8, 1)]
            table_intertops = [(2, 7), (6, 3), (4, 5), (8, 1)]
            #Le 1er du top va en position 2
            #Le 2e du top va en position 7
            #Le 3e du top va en position 6
            #etc ...
            self.set_new_player(level, table_intertops[index_match][0], player1, nb_tops, dossard1, tableau, intertop)
            self.set_new_player(level, table_intertops[index_match][1], player2, nb_tops, dossard2, tableau, intertop)
            if (index_match==0):
                msg_clst = f"1er Intertop: {player1}\n2eme Intertop: {player2}"
            else:
                idx_joueur1 = str((index_match*2)+1)
                idx_joueur2 = str((index_match*2)+2)
                msg_clst = f"{idx_joueur1}eme Intertop: {player1}\n{idx_joueur2}eme Intertop: {player2}"
            tableau["intertops"][str(index_match+1)]["score"] = scores_sets
            self.root.withdraw()  # Masquer la fenêtre principale
            messagebox.showinfo(title="Info", message=msg_clst)
            self.root.deiconify()  # Réafficher la fenêtre principale
        else:
            self.set_new_player(level, index_match, player1, nb_tops, dossard1, tableau, intertop)
            tableau[level][str(index_match)]["score"] = scores_sets
            self.root.withdraw()  # Masquer la fenêtre principale
            messagebox.showinfo(title="Info", message=f"{player1} vainqueur")
            self.root.deiconify()  # Réafficher la fenêtre principale
        
        # Remettre le focus sur la fenêtre "Détails du Match"
        match_window.focus_set()        
        
        # Redessiner le tableau après avoir trouvé un vainqueur
        if tableau == self.tableauOK:
            self.draw_bracket(self.tableauOK, self.tourInitial, self.nb_tops * 4, table_name="OK", canvas=self.ok_canvas)
        elif tableau == self.tableauKO:
            self.draw_bracket(self.tableauKO, self.tourInitial, 0, table_name="KO", canvas=self.ko_canvas)
        
    def handle_enter(self, event, entries, player1, player2, level, index_match, tableau, nb_tops, intertop):
        """Gère la validation des scores et détermine le vainqueur."""
        sets_player1 = 0
        sets_player2 = 0
        #mode = self.sets_mode_global.get()
        self.results_modified = [True]
        mode = self.sets_mode_global.get()
        

        # Navigation entre champs
        current_widget = event.widget
        current_index = entries.index(current_widget)

        # Vérifier que la case précédente est remplie, sauf pour la première case
        if current_widget.get() == "":
            current_widget.insert(0, "0")  # Considérer vide comme 0
        elif current_widget.get() == "-":
            current_widget.insert(0, "-0")  # Considérer "-" comme "-0"
            
        if current_index < len(entries):
            # Validation et calcul des scores
            scores_sets =  []
            for entry in entries:
                value = entry.get()
                if value.startswith("-"):
                    sets_player2 += 1
                elif value.isdigit():
                    sets_player1 += 1
                scores_sets.append(value)
                
               
            if (intertop):
                dossard1 = tableau["intertops"][str(index_match+1)].get("dossard_joueur1", "Unknown")
                dossard2 = tableau["intertops"][str(index_match+1)].get("dossard_joueur2", "Unknown")
            else:
                dossard1 = tableau.get(level, {}).get(str(index_match), {}).get("dossard_joueur1", "Unknown")   
                dossard2 = tableau.get(level, {}).get(str(index_match), {}).get("dossard_joueur2", "Unknown")
                
            match_window = event.widget.master  # Référence à la fenêtre "Détails du Match"
            
            if ( (sets_player1 == 2 and mode == "2 sets") or
                (sets_player1 == 3 and mode == "3 sets") ):                        
                self.gestion_vainqueur(player1, player2, dossard1, dossard2, level, index_match, tableau, nb_tops, intertop, scores_sets, match_window)   
                #current_widget.focus_set()
                #event.widget.master.destroy()  # Fermer la fenêtre "Détails"
                #self.restore_focus_to_main_window()  # Remettre le focus sur la fenêtre principale
            elif ( (sets_player2 == 2 and mode == "2 sets") or
                (sets_player2 == 3 and mode == "3 sets") ):                                          
                self.gestion_vainqueur(player2, player1, dossard2, dossard1, level, index_match, tableau, nb_tops, intertop, scores_sets, match_window)
                #current_widget.focus_set()
                #event.widget.master.destroy()  # Fermer la fenêtre "Détails"
                #self.restore_focus_to_main_window()  # Remettre le focus sur la fenêtre principale
            else:
                next_widget = entries[current_index + 1]  # Passer au champ suivant
                next_widget.focus_set()                    
    
    def update_set_visibility(self, current_mode, set_entries, window=None):
        new_mode = self.sets_mode_global.get()
        #update_set_visibility(new_mode, set_entries)
        confirmation_message = "Les données actuelles n'ont pas été sauvegardées. Voulez-vous continuer et effacer les données ?"
        success = handle_mode_change(
            current_mode=self.sets_mode_global.get(),
            new_mode=new_mode,
            result_entries=[set_entries],
            results_modified=self.results_modified
        )
        if success:
            self.results_modified = [False]
            self.previous_mode = new_mode
        else:
            self.sets_mode_global.set(self.previous_mode)
        if window != None:
            window.focus_set()
        update_set_visibility(self.previous_mode, set_entries)
           
    def save_results(self, window):
        """Enregistre les résultats saisis dans un fichier JSON."""
        file_name = f"Tableau_{self.category}_tour{self.round_number}.json"
     
        # Charger les données existantes si le fichier existe
        data = {
            "category": self.category,
            "round": self.round_number,
            "tourInitial": self.tourInitial,
            "nbTops": self.nb_tops,
            "nbPoules": self.nb_poules,
            "Tableau OK": self.tableauOK,
            "Tableau KO": self.tableauKO
        }
        # Enregistrer dans le fichier
        
        resp = save_tableau(self.fichier_Tableau, data)
        if (resp):
            window.destroy()
            self.results_modified = [False]
            self.load_tournament(self.current_table)
        
    def open_match_window(self, player1, player2, string_level, index_match, tableau, nb_tops, intertop=False):
        # Créer une nouvelle fenêtre
        match_window = tk.Toplevel(self.root)
        match_window.title("Détails du Match")
        
        # Déterminer la fenêtre parent
        if self.current_table == "Tournoi OK":
            parent_window = self.ok_window
        elif self.current_table == "Tournoi KO":
            parent_window = self.ko_window
        elif self.current_table == "InterTops":
            parent_window = self.intertops_window

        # Associer la fenêtre "Détails du Match" à la fenêtre parent
        #parent_window = self.ok_window if self.current_table == "Tournoi OK" else self.ko_window
        match_window.transient(parent_window)  # Associer à la fenêtre parent
        match_window.grab_set()  # Rendre la fenêtre modale

        # Ajouter les noms des joueurs
        tk.Label(match_window, text=f"{player1} vs {player2}", font=("Arial", 14)).pack(pady=10)

        # Radio buttons pour 2 sets ou 3 sets
        
        self.sets_mode_global = tk.StringVar(value=self.previous_mode)
        ttk.Radiobutton(match_window, text="2 sets", variable=self.sets_mode_global, value="2 sets", 
            command=lambda: self.update_set_visibility(self.sets_mode_global.get(), set_entries, match_window)).pack()

        ttk.Radiobutton(match_window, text="3 sets", variable=self.sets_mode_global, value="3 sets", 
            command=lambda: self.update_set_visibility(self.sets_mode_global.get(), set_entries, match_window)).pack()

        def on_focus_in(event, current_index):
            if current_index < len(set_entries) - 1:
                for idx in range(0, current_index):
                    if (set_entries[idx].get() == ""):
                        set_entries[idx].focus_set()
                        break 
   
        # Champs pour les scores des sets
        set_entries = []
        for i in range(5):
            row = tk.Frame(match_window)
            row.pack(pady=5)
            tk.Label(row, text=f"Set {i + 1}:", font=("Arial", 10)).pack(side="left")
            #entry = tk.Entry(row, width=5, validate="key", validatecommand=(self.root.register(self.validate_set_input), '%d', '%P'))
            entry = tk.Entry(row, width=5, validate="key",
                                  validatecommand=(self.root.register(lambda action, value: validate_input_set(action, value, self.results_modified)
                            ), '%d', '%P'))
            entry.pack(side="left")
            entry.bind("<Return>", lambda event, e=set_entries, p1=player1, p2=player2, lvl=string_level, idx=index_match, nbt=nb_tops: self.handle_enter(event, e, p1, p2, lvl, idx, tableau, nbt, intertop))
            #entry.bind("<KeyRelease>", lambda e, idx=i: validate_next_set(idx))
            entry.bind("<FocusIn>", lambda e, idx=i: on_focus_in(e, idx))
            set_entries.append(entry)

        button_frame = ttk.Frame(match_window)
        button_frame.pack(fill="x", pady=10)

        # Bouton pour fermer la fenêtre
        quit_button = ttk.Button(button_frame, text="Fermer", command=lambda w=match_window: self.quit_match(w))
        quit_button.pack(side="right", padx=5)

        # Bouton pour valider
        validate_button = ttk.Button(button_frame, text="Valider", command=lambda w=match_window: self.save_results(w))
        validate_button.pack(side="left", padx=5)
        
        # Mettre à jour la visibilité initiale des sets
        self.update_set_visibility("2 sets", set_entries)
        
        # Forcer un redimensionnement
        root.update_idletasks()  # Mettre à jour l'interface
        
    def draw_intertops(self, tableau, canvas):
        index_match = 0
        canvas.delete("all")
            
        # Ajuster la hauteur dynamique du canevas
        canvas_height = max(600, 400 + 8 * 70 * 2)
        #self.canvas.config(scrollregion=(0, 0, 1200, canvas_height))
        start_x = 20
        player_spacing = canvas_height // 8 // 4
        gap = 10  # Espace entre les cadres
        longueur_rectangle = 180
        longueur_fleche = 100
        
        for idx in range (0, 4):
            joueur1 = ""
            joueur2 = ""
            y1_top = idx * 2 * (player_spacing + gap) + 10
            y1_bottom =  y1_top + player_spacing
            mid_y1 = (y1_top + y1_bottom) / 2
            match_nb = str(idx + 1)
            rect_id_1 = canvas.create_rectangle(start_x, y1_top, start_x + longueur_rectangle, y1_bottom, outline="black", width=1)
            
            if (match_nb in tableau['intertops'] and  "joueur1" in tableau['intertops'][match_nb]):
                joueur1 = tableau['intertops'][match_nb]['joueur1']
                dossard1 = tableau['intertops'][match_nb].get('dossard_joueur1', "")
                canvas.create_text(start_x + 10, y1_top + 15, text=joueur1, anchor="w", font=("Arial", 10))

            y2_top = y1_bottom + 10
            y2_bottom = y2_top + player_spacing
            
            mid_y2 = (y2_top + y2_bottom) // 2
            rect_id_2 = canvas.create_rectangle(start_x, y2_top, start_x + longueur_rectangle, y2_bottom, outline="black", width=1)
            
            if (match_nb in tableau['intertops'] and  "joueur2" in tableau['intertops'][match_nb]):
                joueur2 = tableau['intertops'][match_nb]['joueur2']
                dossard2 = tableau['intertops'][match_nb].get('dossard_joueur2', "")
                canvas.create_text(start_x + 10, y2_top + 15, text=joueur2, anchor="w", font=("Arial", 10))
                
                
            def on_enter(event, dossard_text, joueur_text):
                """Crée et affiche la fenêtre Toplevel quand la souris entre dans le rectangle."""
                if not dossard_text.strip() and not joueur_text.strip():
                    # Si vide => pas d'infobulle
                    return
    
                self._tooltip_window = tk.Toplevel(canvas)
                self._tooltip_window.wm_overrideredirect(True)  # supprime la barre d’entête
                self._tooltip_window.geometry(f"+{event.x_root+10}+{event.y_root+10}")

                doss = f"Dossard: {dossard_text}" if dossard_text else "Dossard inconnu"
                label = tk.Label(self._tooltip_window, text=doss,
                     background="#ffffe0", relief="solid", borderwidth=1)
                label.pack()

            def on_leave(event):
                """Détruit la fenêtre Toplevel quand la souris sort du rectangle."""
                if getattr(self, "_tooltip_window", None) is not None:
                    self._tooltip_window.destroy()
                    self._tooltip_window = None

            # Lier les événements <Enter> et <Leave> au rectangle
            canvas.tag_bind(rect_id_1, "<Enter>", 
                lambda e, doss=dossard1, jr=joueur1: on_enter(e, doss, jr))
            canvas.tag_bind(rect_id_1, "<Leave>", on_leave)
            
            canvas.tag_bind(rect_id_2, "<Enter>", 
                lambda e, doss=dossard2, jr=joueur2: on_enter(e, doss, jr))
            canvas.tag_bind(rect_id_2, "<Leave>", on_leave)
            
            mid_y = (mid_y1 + mid_y2) // 2
            
            button_x = start_x + longueur_rectangle + 20
            button_y = mid_y
            state = "normal"
            intertop = True
            btn = tk.Button(canvas, text="Détails", state=state, command=lambda p1=joueur1, p2=joueur2, lvl="InterTops", match_index=idx: self.open_match_window(p1, p2, lvl, match_index, tableau, self.nb_tops, intertop))
            canvas.create_window(button_x, button_y, window=btn)
            
            checkbox_x = button_x + 50  # Décalage pour la case
            var = tk.BooleanVar()
            
            if "intertops" in tableau:
                stored_value = tableau["intertops"].get(str(idx+1), {}).get("checkbox_state", False)
                var.set(stored_value)
            else:
                var.set(False)
    
            def on_checkbox_change_intertops(var, match_index, tableau):
                """Callback pour mettre à jour le 'checkbox_state' d’un match Intertops."""
                # On s’assure que la structure existe :
                self.results_modified[0] = True
                if "intertops" not in tableau:
                    tableau["intertops"] = {}
                if str(match_index) not in tableau["intertops"]:
                    tableau["intertops"][str(match_index)] = {}
                tableau["intertops"][str(match_index)]["checkbox_state"] = var.get()
            
            chk = tk.Checkbutton(
                canvas, variable=var,
                command=lambda v=var, m=idx+1: on_checkbox_change_intertops(v, m, tableau)
                )
            canvas.create_window(checkbox_x, button_y, window=chk)

            # Mémoriser cette variable pour la sauvegarde dans save_results
            self.checkbox_vars[("intertops", idx+1)] = var

            #checkbox_x = button_x + 50  # Ajustez pour positionner la case
            #var = tk.BooleanVar()

            # Récupérer la valeur stockée s’il y en a une (checkbox_state) :
            #stored_value = tableau.get(string_level, {}).get(str(idx), {}).get("checkbox_state", False)
            #var.set(stored_value)

            #chk = tk.Checkbutton(self.canvas, variable=var)
            #self.canvas.create_window(checkbox_x, button_y, window=chk)

            # Mémoriser cette variable pour la sauvegarde
            #self.checkbox_vars[(string_level, idx)] = var

    def print_bracket_to_pdf(self, canvas):
        import os
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas as pdf_canvas
        from PIL import Image
        
        pdf_directory = lire_pdf_directory()
        os.makedirs(pdf_directory, exist_ok=True)  # Créer le dossier s'il n'existe pas

        #table_type = "OK" if "OK" in self.current_table else "KO"
        table_type = "OK" if canvas == self.ok_canvas else "KO"

        # Définir le nom du fichier de sortie
        pdf_filename = os.path.join(pdf_directory, f"Tableau_{table_type}_{self.category}_Tour{self.round_number}.pdf")
        
        # 1) Déterminer la zone totale du Canvas
        bbox = canvas.bbox("all")

        # ⚠️ Vérifier si `bbox` est `None`
        if bbox is None:
            messagebox.showerror("Impression", "Rien à imprimer ou dimensions invalides.")
            return

        left, top, right, bottom = bbox
        width_canvas = right - left
        height_canvas = bottom - top

        if width_canvas <= 0 or height_canvas <= 0:
            messagebox.showerror("Impression", "Rien à imprimer ou dimensions invalides.")
            return
    
        # 2) Exporter le canvas en PostScript
        ps_filename = os.path.join(pdf_directory, "temp_bracket.ps")
        try:
            canvas.postscript(file=ps_filename, colormode="color", x=left, y=top, width=width_canvas, height=height_canvas)
        except Exception as e:
            messagebox.showerror("Impression", f"Échec de l'export PostScript : {e}")
            return

        # 3) Convertir PostScript -> Image (PNG) avec une meilleure résolution
        png_filename = os.path.join(pdf_directory, "temp_bracket.png")
        try:
            img = Image.open(ps_filename)
            
            # Convertir l'image en mode RGB si nécessaire
            if img.mode != "RGB":
                img = img.convert("RGB")
            print("Image mode after conversion:", img.mode)

            # Augmenter la résolution pour meilleure lisibilité
            scale_factor = 4  # Multiplier la taille par 2 pour plus de netteté
            new_width = int(img.width * scale_factor)
            new_height = int(img.height * scale_factor)

            #img = img.resize((new_width, new_height), RESAMPLE_METHOD)
            img.resize((new_width, new_height), Image.Resampling.LANCZOS)
            img.save(png_filename, "PNG", dpi=(300, 300))
            
            img.close()
        except Exception as e:
            messagebox.showerror("Impression", f"Échec de la conversion en PNG : {e}")
            return

        # 4) Générer le PDF avec la bonne mise en page
        try:
            #pdf_canvas = canvas.Canvas(pdf_filename, pagesize=A4)
            pdf = pdf_canvas.Canvas(pdf_filename, pagesize=A4)
            pdf_width, pdf_height = A4  # Dimensions d'une page A4

            # Déterminer l'échelle pour que le tableau rentre sur une page
            scale_x = pdf_width / new_width
            scale_y = pdf_height / new_height
            scale = min(scale_x, scale_y)  # Prendre le plus petit facteur pour tout afficher

            # Calculer les marges pour centrer l'image
            final_width = new_width * scale
            final_height = new_height * scale
            margin_x = (pdf_width - final_width) / 2
            margin_y = (pdf_height - final_height) / 2

            # Ajouter l'image au PDF
            pdf.drawImage(png_filename, margin_x, margin_y, width=final_width, height=final_height)
        
            pdf.showPage()
            pdf.save()

            messagebox.showinfo("Impression", f"PDF généré avec succès : {pdf_filename}")

        except Exception as e:
            messagebox.showerror("Impression", f"Échec de la génération du PDF : {e}")
            return

        # Optionnel : Supprimer les fichiers temporaires
        os.remove(ps_filename)
        os.remove(png_filename)
        
    #def draw_bracket(self, tableau, firstTour, nb_players_top, table_name):
    def draw_bracket(self, tableau, firstTour, nb_players_top, table_name, canvas):    
        #self.canvas.delete("all")
        canvas.delete("all")
        
        nb_players = 0
        if self.nb_poules == 2:
            nb_players = 2
        elif self.nb_poules <= 4:
            nb_players = 4       
        elif self.nb_poules <= 8:
            nb_players = 8
        elif self.nb_poules <= 16:
            nb_players = 16
        elif self.nb_poules <= 32:
            nb_players = 32
        print(f"nb_players:{nb_players}, firstTour:{firstTour}")

        # Ajuster la hauteur dynamique du canevas
        canvas_height = max(600, 300 + nb_players * 50 * 2)
        canvas.config(scrollregion=(0, 0, 1200, canvas_height))

        player_spacing = canvas_height // nb_players // 6
        gap = 5
        next_round_positions = []

        header_text = f"Tableau {self.category} {table_name} - Tour {self.round_number}"
        #canvas.create_text(500, 50, text=f"{table_name} - Tour {self.round_number}", font=("Arial", 16, "bold"))
        canvas.create_text(500, 50, text=header_text, font=("Arial", 16, "bold"))
        
        def on_click_print():
            self.print_bracket_to_pdf(canvas)
        
        # Créer des cadres pour les joueurs et connecter les lignes                    
        index_match = 0
        for idx in range (0, nb_players*2):
                #y1_top = idx * 2 * (player_spacing + gap) + 10
                y1_top = idx * 2 * (player_spacing + gap) + 60
                y1_bottom =  y1_top + player_spacing
                mid_y = (y1_top + y1_bottom) / 2
                index_match = 0
                next_round_positions.append(mid_y)
            
        new_positions = []
        last_match_y = 0
        start_x = 20
        
        currentTour = firstTour
        
        nb_matches = self.tourInitial*2
            
        string_level = f"1/{currentTour}"
        
        if (string_level in tableau):
            matchesTour = tableau[string_level]
            
        while (nb_matches >=1):
            new_positions = []
            index_player = 1    
            longueur_rectangle = 130  # Réduction de la largeur des rectangles
            longueur_fleche = 50  # Réduction de la longueur des flèches
            
            level_found_in_matches = False
            string_level = f"1/{currentTour}"
            
            if currentTour == 1:
                label_text = "Finale"
            elif currentTour == 0:
                label_text = "Vainqueur"
            else:
                label_text = f"1/{currentTour}"
                
            canvas.create_text(
                start_x + (longueur_rectangle / 2),   # Coordonnée X (le centre de la colonne)
                20,                                   # Coordonnée Y fixe (en haut)
                text=label_text,
                anchor="center",
                font=("Arial", 10, "bold")
            )
            
            if (string_level in tableau):
                matches_list = tableau[string_level]
                level_found_in_matches = True
            idx = 1
            j1 = ""
            j2 = ""
            index_player = 1
            for round_mid_y in next_round_positions:
                string_level = f"1/{currentTour}"
                if (level_found_in_matches):
                    matches_list = tableau[string_level]
                else:
                    matches_list = []
                key= str(idx)

                def display_entree(round_mid_y, dossard_text, text, level_str, match_idx):
                    y_top = (round_mid_y - player_spacing//2)
                    y_bottom = y_top + player_spacing
                    mid_y = (y_top + y_bottom) // 2
                    fillColor = "white"
                    if (text == "---"):
                        fillColor = "grey"
                        text = ""
                    rect_id = canvas.create_rectangle(start_x, y_top, start_x + longueur_rectangle, y_bottom, outline="black", width=1, fill=fillColor)
                    #if (text != ""):
                    text_id = canvas.create_text(start_x + 3, y_top + 8, text=text, anchor="w", font=("Arial", 9))  # Réduction de la taille du texte
                        
                         

                        
                    def on_enter(event):
                        """Crée et affiche la fenêtre Toplevel quand la souris entre dans le rectangle."""
                        # On se souvient du tip dans un attribut (ex: self._tooltip_window) 
                        # pour pouvoir le fermer plus tard.
                        if not dossard_text.strip() and not text.strip():
                            # Si vide => on ne fait rien (pas de tooltip)
                            return
                        
                        self._tooltip_window = tk.Toplevel(canvas)
                        self._tooltip_window.wm_overrideredirect(True)  # Supprime la barre d'entête
                        # On positionne près du curseur : event.x_root, event.y_root
                        self._tooltip_window.geometry(f"+{event.x_root+10}+{event.y_root+10}")

                        # On y place un label avec le dossard
                        doss = f"Dossard: {dossard_text}" if dossard_text else "Dossard inconnu"
                        label = tk.Label(self._tooltip_window, text=doss, 
                         background="#ffffe0", relief="solid", borderwidth=1)
                        label.pack()

                    def on_leave(event):
                        """Détruit la fenêtre Toplevel quand la souris sort du rectangle."""
                        if getattr(self, "_tooltip_window", None) is not None:
                            self._tooltip_window.destroy()
                            self._tooltip_window = None
                    
                    '''checkbox_x = start_x + longueur_rectangle + 5
                    checkbox_y = mid_y
                    create_checkbox(checkbox_x, checkbox_y, string_level, idx)
                    '''
                    '''group_tag = f"{string_level}_{idx}"  # un identifiant de votre choix
                    self.canvas.itemconfig(rect_id, tags=(group_tag,))
                    self.canvas.itemconfig(text_id, tags=(group_tag,))

                    self.canvas.tag_bind(group_tag, "<Enter>", lambda e: on_enter(e, dossard_text, joueur_text))
                    self.canvas.tag_bind(group_tag, "<Leave>", on_leave)
                    '''
                    canvas.tag_bind(rect_id, "<Enter>", on_enter)
                    canvas.tag_bind(rect_id, "<Leave>", on_leave)
                    
                    return y_top, y_bottom, mid_y
                
                def on_checkbox_change(var, level, match_index, tableau):
                    self.results_modified[0] = True
                    if level  in tableau and str(match_index) in tableau[level]:
                        cbox_string = f"cbx_match_{match_index}"
                        tableau[level][str(match_index)]["checkbox_state"] = var.get()

                if (index_player == 1):
                    if key in matches_list and "joueur1" in matches_list[key]:
                        j1 = matches_list[str(idx)]['joueur1']
                        d1 = matches_list[key].get('dossard_joueur1', '')
                    else:
                        j1, d1 = '', ''
                    y1_top,y1_bottom,mid_y1 = display_entree(round_mid_y, d1, j1, string_level, idx)
                    if (nb_matches == 2*nb_players_top):
                        new_positions.append(round_mid_y)
                    index_player = index_player + 1
                elif (index_player == 2):
                    if key in matches_list and "joueur2" in matches_list[key]:
                        j2 = matches_list[str(idx)]['joueur2']
                        d2 = matches_list[key].get('dossard_joueur2', '')
                    else:
                        j2, d2 = '', ''
                    y2_top,y2_bottom,mid_y2 = display_entree(round_mid_y, d2, j2, string_level, idx)       
                    mid_y = (mid_y1 + mid_y2) // 2
                    index_player = 1
                    canvas.create_line(start_x + longueur_rectangle, mid_y1, start_x + longueur_rectangle + longueur_fleche, mid_y, fill="black")
                    canvas.create_line(start_x + longueur_rectangle, mid_y2, start_x + longueur_rectangle + longueur_fleche, mid_y, fill="black")
                    
                    if (nb_matches == 2*nb_players_top):
                        new_positions.append(round_mid_y)
                    else:
                        new_positions.append(mid_y)
                        
                    if (nb_matches == 1):
                        last_match_y = mid_y
                        
                    button_x = start_x + 10
                    button_y = mid_y
                    
                    state = "normal"
                    if (self.debugStatus=="False" and (j1 == "" or j1 == "---" or j2 == "" or j2 == "---") ):
                        state="disabled"  
                    btn = tk.Button(canvas, text="Détails", state = state, command=lambda p1=j1, p2=j2, lvl=string_level, match_index=idx, nbt=nb_players_top: self.open_match_window(p1, p2, lvl, match_index, tableau, nbt))
                    canvas.create_window(button_x, button_y, window=btn)
                    
                    checkbox_x = button_x + 50  # Ajustez pour positionner la case
                    var = tk.BooleanVar()

                    # Récupérer la valeur stockée s’il y en a une (checkbox_state) :
                    cbox_string = f"cbx_match_{idx}"
                    if string_level not in tableau or str(idx) not in tableau[string_level] or "checkbox_state" not in tableau[string_level][str(idx)]:
                        var.set(False)
                    else:
                        var.set(tableau[string_level][str(idx)]["checkbox_state"])

                    chk = tk.Checkbutton(canvas, variable=var, command=lambda elt=var, lvl=string_level, match_index=idx: on_checkbox_change(elt, lvl, match_index, tableau))
                    chk.configure(state=state)
                    canvas.create_window(checkbox_x, button_y, window=chk)

                    # Mémoriser cette variable pour la sauvegarde
                    self.checkbox_vars[(string_level, idx)] = var
                    
                    
                    j1 = ""
                    j2 = ""
                        
                    idx = idx + 1
            start_x = start_x + longueur_rectangle + longueur_fleche
            currentTour = currentTour // 2
            next_round_positions = new_positions            
            if (nb_matches == 2*nb_players_top):
                nb_players_top = 0
            else:
                nb_matches = nb_matches / 2

    def read_config_files(self):
        self.fichier_Tableau = f"Tableau_{self.category}_tour{self.round_number}.json"
        print(f"FICHIER TABLEAU: {self.fichier_Tableau}")
        #self.data = load_all_tableau(self.fichier_Tableau, True)
        
    def restore_focus_to_main_window(self):
        """Remet le focus sur la fenêtre principale (Tableau OK ou Tableau KO)."""
        print ("current_table:", self.current_table)
        if self.current_table == "Tournoi OK" and self.ok_window:
            self.ok_window.focus_set()
        elif self.current_table == "Tournoi KO" and self.ko_window:
            self.ko_window.focus_set()
    
if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage : python Saisie-Tournoi-V1.5.py <categorie> <tour>")
        sys.exit(1)

    category_arg = sys.argv[1]
    round_arg = sys.argv[2]    
        
    root = tk.Tk()
    app = TournamentBracket(root, category_arg, round_arg)
    root.mainloop()
