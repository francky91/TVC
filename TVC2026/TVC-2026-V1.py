import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import subprocess
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from modules.utils import charger_config, lire_prg_config, lire_excel_name
from modules.github_access import (
    get_github_url,
    get_github_headers,
    list_github_files,
    get_github_file,
)
import os
import shutil
import configparser

# Dictionnaires pour suivre les fenêtres ouvertes par catégorie
opened_windows_poules = {}
opened_windows_tableaux = {}
opened_windows_repartition = {}
opened_windows_intertops = {}

def check_all_processes():
    """Vérifie l'état de tous les processus ouverts."""
    for category, process in list(opened_windows_poules.items()):
        if process.poll() is not None:  # Si le processus est terminé
            del opened_windows_poules[category]
            print(f"Application 'Poules' fermée pour la catégorie '{category}'.")

    for category, process in list(opened_windows_tableaux.items()):
        if process.poll() is not None:  # Si le processus est terminé
            del opened_windows_tableaux[category]
            print(f"Application 'Tableaux' fermée pour la catégorie '{category}'.")
            
    for category, process in list(opened_windows_repartition.items()):
        if process.poll() is not None:  # Si le processus est terminé
            del opened_windows_repartition[category]
            print(f"Application 'Repartition' fermée pour la catégorie '{category}'.")

    # Replanifier la vérification
    root.after(1000, check_all_processes)

def open_app(category, round_number, app_type, prg_name):
    """Ouvre une application spécifique (poules ou tableaux) avec les paramètres donnés."""
    if app_type == "poules":
        if category in opened_windows_poules:
            messagebox.showwarning("Attention", f"L'application pour la catégorie '{category}' (Poules) est déjà ouverte.")
            return
        #app_name = 'saisie-resultats-poules-V2.3.py'
        opened_windows = opened_windows_poules
    elif app_type == "tableaux":
        if category in opened_windows_tableaux:
            messagebox.showwarning("Attention", f"L'application pour la catégorie '{category}' (Tableaux) est déjà ouverte.")
            return
        #app_name = 'Saisie-Tournoi-V1.py'  # Remplacez par le nom réel de votre application pour les tableaux
        opened_windows = opened_windows_tableaux
    elif app_type == "repartition":
        if category in opened_windows_repartition:
            messagebox.showwarning("Attention", f"L'application pour la catégorie '{category}' (Repartition) est déjà ouverte.")
            return
        #app_name = 'Saisie-Tournoi-V1.py'  # Remplacez par le nom réel de votre application pour les tableaux
        opened_windows = opened_windows_repartition
    else:
        raise ValueError("Type d'application non valide.")

    try:
        # Lancer le processus pour l'application correspondante
        print (f"prg_name: {prg_name}")
        process = subprocess.Popen(['python', prg_name, category, str(round_number)])
        opened_windows[category] = process
    except Exception as e:
        messagebox.showerror("Erreur", f"Impossible d'ouvrir l'application '{app_type}' : {e}")

def on_open_button_click(action, prg_name):
    """Gère les clics sur les boutons 'Ouvrir Poules' et 'Ouvrir Tableau'."""
    selected_category = category_var.get()
    selected_round = round_var.get()

    if not selected_category:
        messagebox.showerror("Erreur", "Veuillez sélectionner une catégorie")
        return
    elif not selected_round and action != "repartition":
        messagebox.showerror("Erreur", "Veuillez sélectionner un numéro de tour.")
        return

    if action == "tableaux":
        # Vérifier si le fichier JSON existe
        json_filename = f"Tableau_{selected_category}_tour{selected_round}.json"
        json_path = os.path.join(JsonDir, json_filename)
        if not os.path.exists(json_path):
            messagebox.showerror("Erreur", f"Veuillez rentrer au moins un résultat de poule")
            return
        
    if action == "poules":
        open_app(selected_category, selected_round, "poules", prg_name)
    elif action == "tableaux":
        open_app(selected_category, selected_round, "tableaux", prg_name)
    elif action == "repartition":
        open_app(selected_category, selected_round, "repartition", prg_name)
    
def update_round_combobox(*args):  # AJOUT
    """
    Met à jour la combobox round_combobox en fonction des fichiers JSON
    existants pour la catégorie sélectionnée.
    """
    selected_category = category_var.get()
    if not selected_category:
        return

    available_rounds = []
    # Parcourez les tours possibles (1, 2, 3, 4) ou une autre plage si vous le souhaitez
    for r in [1, 2, 3, 4]:
        json_filename = f"{selected_category}_tour{r}.json"
        json_path = os.path.join(JsonDir, json_filename)  # Utilisation de JsonDir
        if os.path.exists(json_path):
            available_rounds.append(str(r))

    # Mettre à jour les valeurs de la combobox
    round_combobox['values'] = available_rounds
    
    if available_rounds:
        round_combobox.set(available_rounds[0])
        open_poules_button.config(state="normal")  # Activer le bouton "Ouvrir Poules"
        open_tableau_button.config(state="normal")  # Activer le bouton "Ouvrir Tableau"
    else:
        round_combobox.set('')  # aucune sélection possible
        open_poules_button.config(state="disabled")  # Désactiver le bouton "Ouvrir Poules"
        open_tableau_button.config(state="disabled")  # Désactiver le bouton "Ouvrir Tableau"
    
    # Sélectionner automatiquement le premier tour existant si disponible
    if available_rounds:
        round_combobox.set(available_rounds[0])
    else:
        round_combobox.set('')  # aucune sélection possible
        
def refresh_window():
    """Actualise la fenêtre principale."""
    # Vous pouvez ajouter ici le code pour actualiser les données ou l'interface
    update_round_combobox()

def _ask_file_choice(file_names):
    """Ouvre une petite fenêtre pour choisir un fichier et retourne son nom, ou None."""
    dialog = tk.Toplevel(root)
    dialog.title("Sélectionner un fichier")
    dialog.geometry("400x300")
    dialog.transient(root)
    dialog.grab_set()

    selected_file = [None]

    frame = ttk.Frame(dialog, padding=10)
    frame.pack(fill="both", expand=True)

    ttk.Label(frame, text="Sélectionnez un fichier :").pack(pady=5)

    listbox = tk.Listbox(frame, height=10)
    listbox.pack(fill="both", expand=True, pady=5)

    # On trie les noms et on se base uniquement sur cette liste
    sorted_files = sorted(file_names)
    for name in sorted_files:
        listbox.insert(tk.END, name)

    def on_select():
        selection = listbox.curselection()
        if selection:
            selected_file[0] = sorted_files[selection[0]]
            dialog.destroy()

    def on_cancel():
        dialog.destroy()

    button_frame = ttk.Frame(frame)
    button_frame.pack(pady=5)

    ttk.Button(button_frame, text="Sélectionner", command=on_select).pack(side="left", padx=5)
    ttk.Button(button_frame, text="Annuler", command=on_cancel).pack(side="left", padx=5)

    listbox.bind("<Double-1>", lambda e: on_select())

    dialog.wait_window()
    return selected_file[0]


def load_file():
    """Point d'entrée du bouton 'Charger Fichier'."""
    try:
        config_file = "config.ini"

        github_url = get_github_url(config_file)
        if not github_url:
            return

        headers = get_github_headers()
        if not headers:
            return

        file_names = list_github_files(github_url, headers)
        if not file_names:
            return

        file_name = _ask_file_choice(file_names)
        if not file_name:
            return

        # Lecture du répertoire Excel depuis la config
        config = configparser.ConfigParser()
        config.read(config_file)
        try:
            directory_excel = config.get("Paths", "directoryExcel")
        except Exception:
            messagebox.showerror("Erreur", "directoryExcel non trouvé dans config.ini")
            return

        os.makedirs(directory_excel, exist_ok=True)
        local_file_path = os.path.join(directory_excel, file_name)
        file_exists = os.path.exists(local_file_path)

        should_replace = False
        if file_exists:
            response = messagebox.askyesnocancel(
                "Fichier existant",
                f"Le fichier '{file_name}' existe déjà.\n\n"
                f"Oui : Remplacer (backup du fichier précédent)\n"
                f"Non : Ne rien faire\n"
                f"Annuler : Annuler l'opération"
            )

            if response is None:  # Annuler
                return
            elif response:  # Oui - remplacer
                should_replace = True
            else:  # Non - ne rien faire
                return

        # Télécharger le contenu du fichier depuis GitHub
        file_url = f"{github_url}/{file_name}"
        file_content, _ = get_github_file(file_url, headers)
        if file_content is None:
            return

        # Sauvegarde de backup si besoin
        if file_exists and should_replace:
            backup_path = f"{local_file_path}_backup.xlsx"
            shutil.copy2(local_file_path, backup_path)
            messagebox.showinfo("Backup", f"Fichier sauvegardé : {backup_path}")

        # Écriture du nouveau fichier
        with open(local_file_path, "wb") as f:
            f.write(file_content)

        messagebox.showinfo(
            "Succès",
            f"Fichier '{file_name}' copié avec succès.\n\n"
            f"Veuillez modifier le fichier config.ini et redémarrer l'application."
        )

    except Exception as e:
        messagebox.showerror("Erreur", f"Erreur lors du chargement du fichier : {e}")

def load_category():
    """
    Remplace l'onglet de la catégorie sélectionnée dans le fichier Excel local
    par celui du fichier Excel distant sur GitHub.
    """
    try:
        # 1) Récupérer le chemin du fichier Excel local
        local_excel = lire_excel_name()
        if not os.path.exists(local_excel):
            messagebox.showerror("Erreur", f"Fichier Excel local introuvable :\n{local_excel}")
            return

        # 2) Catégorie sélectionnée
        categorie = category_var.get()
        if not categorie:
            messagebox.showerror("Erreur", "Veuillez sélectionner une catégorie.")
            return

        # 3) URL GitHub et headers
        github_url = get_github_url()
        if not github_url:
            return

        headers = get_github_headers()
        if not headers:
            return

        file_name = os.path.basename(local_excel)
        file_url = f"{github_url}/{file_name}"

        # 4) Vérifier que le fichier Excel existe sur GitHub et récupérer son contenu
        remote_bytes, _ = get_github_file(file_url, headers)
        if remote_bytes is None:
            # get_github_file affiche déjà un message d'erreur
            return

        # 5) Backup du fichier local (for<categorie>)
        base, ext = os.path.splitext(local_excel)
        backup_path = f"{base}_for{categorie}{ext}"
        shutil.copy2(local_excel, backup_path)
        messagebox.showinfo("Backup", f"Fichier local sauvegardé :\n{backup_path}")

        # 6) Charger les classeurs distant et local
        remote_wb = load_workbook(BytesIO(remote_bytes))
        local_wb = load_workbook(local_excel)

        if categorie not in remote_wb.sheetnames:
            messagebox.showerror(
                "Erreur",
                f"L'onglet '{categorie}' est introuvable dans le fichier distant.",
            )
            return

        if categorie not in local_wb.sheetnames:
            messagebox.showerror(
                "Erreur",
                f"L'onglet '{categorie}' est introuvable dans le fichier local.",
            )
            return

        remote_ws = remote_wb[categorie]
        local_ws = local_wb[categorie]

        # 7) Effacer toutes les valeurs de l'onglet local (en laissant la mise en forme/merges)
        for row in local_ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None

        # 8) Recopier toutes les valeurs de l'onglet distant dans l'onglet local
        for row in remote_ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                local_ws[cell.coordinate].value = cell.value

        # 9) Sauvegarder le fichier local
        local_wb.save(local_excel)

        messagebox.showinfo(
            "Succès",
            f"Onglet '{categorie}' mis à jour depuis GitHub dans le fichier :\n{local_excel}",
        )

    except Exception as e:
        messagebox.showerror("Erreur", f"Erreur lors du chargement de la catégorie : {e}")

# Création de la fenêtre principale
root = tk.Tk()
root.title("TVC : Gestion des Poules et Tableaux")

root.geometry("400x400")

logo_img = tk.PhotoImage(file="logo.png")

# Création d’un Label pour afficher le logo
logo_label = tk.Label(root, image=logo_img)
logo_label.pack(pady=10)

# Lire la config
print("Lecture du fichier de configuration 1...")
config_file = "config.ini"  # Assurez-vous que ce fichier existe

PoulesPrg = ""
TableauPrg = ""
RepartitionPrg = ""
JsonDir = "" 
try:
    PoulesPrg,TableauPrg,RepartitionPrg,JsonDir = lire_prg_config(config_file)
    
except FileNotFoundError as e:
    print(e)
except Exception as e:
    print(f"Erreur : {e}")

# Combobox pour les catégories
category_var = tk.StringVar()
categories = ["Poussins", "Benjamins", "Minimes", "Cadets-Juniors", "Feminines"]
category_label = ttk.Label(root, text="Catégorie")
category_label.pack(pady=5)
category_combobox = ttk.Combobox(root, textvariable=category_var, values=categories, state="readonly")
category_combobox.pack(pady=5)
category_combobox.set(categories[0])  # Sélection par défaut

# Combobox pour les numéros de tour
round_var = tk.StringVar()
rounds = ["1", "2", "3", "4"]
round_label = ttk.Label(root, text="Numéro de Tour")
round_label.pack(pady=5)
round_combobox = ttk.Combobox(root, textvariable=round_var, values=rounds, state="readonly")
round_combobox.pack(pady=5)
round_combobox.set(rounds[0])  # Sélection par défaut

# Cadre pour Refresh
button_refresh = ttk.Frame(root)
button_refresh.pack(pady=10)

refresh_button = ttk.Button(button_refresh, text="Refresh", command=refresh_window)
refresh_button.pack(side="left", padx=5)

# Cadre pour les boutons "Ouvrir Poules", "Ouvrir Tableau" et "Repartition Joueurs"
button_frame = ttk.Frame(root)
button_frame.pack(pady=10)

# Bouton "Ouvrir Poules"
open_poules_button = ttk.Button(button_frame, text="Ouvrir Poules", command=lambda: on_open_button_click("poules", PoulesPrg))
open_poules_button.pack(side="left", padx=5)

# Bouton "Ouvrir Tableau"
open_tableau_button = ttk.Button(button_frame, text="Ouvrir Tableau", command=lambda: on_open_button_click("tableaux", TableauPrg))
open_tableau_button.pack(side="left", padx=5)

# Bouton "Repartition Joueurs"
open_repartition_button = ttk.Button(button_frame, text="Repartition Joueurs", command=lambda: on_open_button_click("repartition", RepartitionPrg))
open_repartition_button.pack(side="left", padx=5)

# Cadre pour les boutons de chargement (sous les 3 boutons précédents)
load_frame = ttk.Frame(root)
load_frame.pack(pady=5)

load_file_button = ttk.Button(load_frame, text="Charger Fichier", command=load_file)
load_file_button.pack(side="left", padx=5)

load_category_button = ttk.Button(load_frame, text="Charger Catégorie", command=load_category)
load_category_button.pack(side="left", padx=5)

# Lancer la vérification périodique des processus
root.after(1000, check_all_processes)


category_var.trace_add('write', update_round_combobox)

# MODIF : Initialisation de la catégorie pour déclencher la mise à jour de la combobox
if categories:
    category_var.set(categories[0])  # Cela déclenchera update_round_combobox
    
# Lancer la boucle principale
root.mainloop()
