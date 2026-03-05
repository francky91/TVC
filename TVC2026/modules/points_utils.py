# -*- coding: utf-8 -*-

import json
import os
import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
from openpyxl import load_workbook
from fractions import Fraction
from modules.utils import lire_excel_name
from modules.Tableau import TableauLevels
from modules.models import Player, Match, EMPTY_NAME
from modules.Tournoi import Tournoi
import requests
import base64
from tkinter import simpledialog

def _is_loser(match, player):
    if not match or not match.winner or not player:
        return True
    if (match.player1 and match.player1.key == player.key) or \
       (match.player2 and match.player2.key == player.key):
        return match.winner.key != player.key
    return False

def _is_winner(match, player):
    if not match or not match.winner or not player:
        return False
    return match.winner.key == player.key

def calculate_points(app):
    config = load_points_config_for_app(app)
    points = {}

    def safe_fraction_from_level_string(level_str):
        try:
            parts = str(level_str).split("/")
            if len(parts) == 2:
                denom = int(parts[1])
                if denom == 0:
                    return Fraction(9999, 1)
                return Fraction(1, denom)
            return Fraction(str(level_str))
        except Exception:
            return Fraction(9999, 1)

    # --- KO ---
    ko_levels = app.tournoi.tableauKo.tableauLevels
    base_ko = config["KO"]["base_points"]
    inc_ko = config["KO"]["increment_per_round"]
    bonus_ko = config["KO"]["winner_bonus"]

    ko_levels = sorted(ko_levels, key=lambda lvl: safe_fraction_from_level_string(lvl.level))
    nb_ko_tours = app.tournoi.tableauKo.get_nb_tours()
    for i, level in enumerate(ko_levels):
        for match in level.matches:
            for joueur in (match.player1, match.player2):
                if not joueur or joueur.nom == EMPTY_NAME:
                    continue
                if level.level == "1/0" :
                    points[match.player1] = base_ko + (i-1) * inc_ko + bonus_ko
                elif _is_loser(match, joueur):
                    points[joueur] = base_ko + i * inc_ko
                #elif i == nb_ko_tours - 1 and _is_winner(match, joueur):
                #    points[joueur] = base_ko + i * inc_ko + bonus_ko

    # --- OK ---
    vainqueur_ko_points = base_ko + inc_ko * (nb_ko_tours - 1) + bonus_ko
    ok_levels = app.tournoi.tableauOk.tableauLevels
    rel_ok = config["OK"]["relative_to_KO_winner"]
    inc_ok = config["OK"]["increment_per_round"]
    bonus_ok = config["OK"]["winner_bonus"]

    ok_levels = sorted(ok_levels, key=lambda lvl: safe_fraction_from_level_string(lvl.level))
    nb_ok_tours = len(ok_levels)
    base_ok = vainqueur_ko_points + rel_ok

    for i, level in enumerate(ok_levels):
        for match in level.matches:
            for joueur in (match.player1, match.player2):
                if not joueur or getattr(joueur, 'isEmpty', False):
                    continue
                if level.level == "1/0" :
                    points[match.player1] = base_ok + (i-1) * inc_ok + bonus_ok
                elif _is_loser(match, joueur):
                    points[joueur] = base_ok + i * inc_ok

    return points


def load_points_config_for_app(app):
    config_dir = "Json"
    os.makedirs(config_dir, exist_ok=True)
    config_file = os.path.join(config_dir, f"points_config_{app.tournoi.category}.json")
    if os.path.exists(config_file):
        with open(config_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {"KO": {"base_points": 1, "increment_per_round": 1, "winner_bonus": 2},
            "OK": {"relative_to_KO_winner": 0, "increment_per_round": 1, "winner_bonus": 2}}


def save_points_config_for_app(app, config):
    config_dir = "Json"
    os.makedirs(config_dir, exist_ok=True)
    config_file = os.path.join(config_dir, f"points_config_{app.tournoi.category}.json")
    with open(config_file, 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=4, ensure_ascii=False)


def show_config_points_window(app):
    win = tk.Toplevel(app.root)
    win.title("Configuration des regles de points")

    frame = ttk.Frame(win, padding=10)
    frame.pack(fill="both", expand=True)

    config = load_points_config_for_app(app)

    ttk.Label(frame, text="Tableau KO").grid(row=0, column=0, columnspan=2, pady=5)

    ttk.Label(frame, text="Points de base (1er tour KO)").grid(row=1, column=0, sticky="w")
    base_ko_var = tk.IntVar(value=config["KO"]["base_points"])
    ttk.Entry(frame, textvariable=base_ko_var, width=5).grid(row=1, column=1)

    ttk.Label(frame, text="+ points par tour").grid(row=2, column=0, sticky="w")
    inc_ko_var = tk.IntVar(value=config["KO"]["increment_per_round"])
    ttk.Entry(frame, textvariable=inc_ko_var, width=5).grid(row=2, column=1)

    ttk.Label(frame, text="Bonus vainqueur").grid(row=3, column=0, sticky="w")
    win_ko_var = tk.IntVar(value=config["KO"]["winner_bonus"])
    ttk.Entry(frame, textvariable=win_ko_var, width=5).grid(row=3, column=1)

    ttk.Label(frame, text="Tableau OK").grid(row=4, column=0, columnspan=2, pady=5)

    ttk.Label(frame, text="Points perdants 1er tour (relatif au vainqueur KO)").grid(row=5, column=0, sticky="w")
    rel_ok_var = tk.IntVar(value=config["OK"]["relative_to_KO_winner"])
    ttk.Entry(frame, textvariable=rel_ok_var, width=5).grid(row=5, column=1)

    ttk.Label(frame, text="+ points par tour").grid(row=6, column=0, sticky="w")
    inc_ok_var = tk.IntVar(value=config["OK"]["increment_per_round"])
    ttk.Entry(frame, textvariable=inc_ok_var, width=5).grid(row=6, column=1)

    ttk.Label(frame, text="Bonus vainqueur").grid(row=7, column=0, sticky="w")
    win_ok_var = tk.IntVar(value=config["OK"]["winner_bonus"])
    ttk.Entry(frame, textvariable=win_ok_var, width=5).grid(row=7, column=1)

    def save_and_close():
        new_config = {
            "KO": {
                "base_points": base_ko_var.get(),
                "increment_per_round": inc_ko_var.get(),
                "winner_bonus": win_ko_var.get(),
            },
            "OK": {
                "relative_to_KO_winner": rel_ok_var.get(),
                "increment_per_round": inc_ok_var.get(),
                "winner_bonus": win_ok_var.get(),
            }
        }
        save_points_config_for_app(app, new_config)
        win.destroy()

    ttk.Button(frame, text="Sauvegarder", command=save_and_close).grid(row=8, column=0, columnspan=2, pady=10)


def show_points_window(app):
    points_window = tk.Toplevel(app.root)
    points_window.title("Points des joueurs")

    frame = tk.Frame(points_window)
    frame.pack(fill=tk.BOTH, expand=True)

    columns = ("nom", "prenom", "club", "points")
    tree = ttk.Treeview(frame, columns=columns, show="headings", height=15)
    tree.heading("nom", text="Nom")
    tree.heading("prenom", text="Prenom")
    tree.heading("club", text="Club")
    tree.heading("points", text="Points")

    scrollbar_y = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
    tree.configure(yscrollcommand=scrollbar_y.set)
    scrollbar_x = ttk.Scrollbar(frame, orient=tk.HORIZONTAL, command=tree.xview)
    tree.configure(xscrollcommand=scrollbar_x.set)

    tree.grid(row=0, column=0, sticky="nsew")
    scrollbar_y.grid(row=0, column=1, sticky="ns")
    scrollbar_x.grid(row=1, column=0, sticky="ew")
    frame.grid_rowconfigure(0, weight=1)
    frame.grid_columnconfigure(0, weight=1)

    points = calculate_points(app)
    for joueur, pts in sorted(points.items(), key=lambda x: x[1], reverse=True):
        nom, prenom, club = joueur.nom, joueur.prenom, joueur.club
        tree.insert("", tk.END, values=(nom, prenom, club, str(pts)))

    def save_points():
        new_points = {}
        for item in tree.get_children():
            vals = tree.item(item, "values")
            joueur = f"{vals[0]} {vals[1]} {vals[2]}"
            try:
                pts_val = int(vals[3])
            except Exception:
                pts_val = 0
            new_points[joueur] = pts_val

        with open("players_points.json", "w", encoding="utf-8") as f:
            json.dump(new_points, f, indent=4, ensure_ascii=False)

        save_points_in_excel(app, new_points)
        points_window.destroy()


    def save_points_in_github():
        repo_owner = "francky91"
        repo_name = "TVC"
        branch = "main"

        # nom du fichier excel
        fichier_entree = lire_excel_name()
        file_name = os.path.basename(fichier_entree)

        url = f"https://api.github.com/repos/{repo_owner}/{repo_name}/contents/{file_name}"

        # 🔐 demander token
        token = simpledialog.askstring(
        "Authentification GitHub",
        "Entrez votre Personal Access Token GitHub :",
        show="*"
        )

        if not token:
            return

        headers = {
            "Authorization": f"token {token}",
            "Accept": "application/vnd.github+json"
        }

        # ---------------------------------------------------
        # 1️⃣ Télécharger le fichier actuel depuis GitHub
        # ---------------------------------------------------

        r = requests.get(url, headers=headers)

        if r.status_code != 200:
            messagebox.showerror("Erreur", "Impossible de récupérer le fichier sur GitHub.")
            return

        data = r.json()

        file_sha = data["sha"]
        file_content = base64.b64decode(data["content"])

        # sauver temporairement le fichier
        temp_file = "temp_github_excel.xlsx"

        with open(temp_file, "wb") as f:
            f.write(file_content)

        # ---------------------------------------------------
        # 2️⃣ récupérer les points depuis l'interface
        # ---------------------------------------------------

        new_points = {}

        for item in tree.get_children():
            vals = tree.item(item, "values")
            joueur = f"{vals[0]} {vals[1]} {vals[2]}"

            try:
                pts_val = int(vals[3])
            except:
                pts_val = 0

            new_points[joueur] = pts_val

        # ---------------------------------------------------
        # 3️⃣ modifier uniquement la feuille concernée
        # ---------------------------------------------------
        try:
            wb = load_workbook(temp_file)
            feuille = wb[app.tournoi.category]

            wb_local = load_workbook(fichier_entree)
            feuille_local = wb_local[app.tournoi.category]

            tour = int(app.tournoi.tour)
            col_presence = 8 + (tour - 1) * 2
            for row in range(8, 300):
                val = feuille_local.cell(row=row, column=col_presence).value
                if val in ("x", "X"):
                    feuille.cell(row=row, column=col_presence).value = val

            tour = int(app.tournoi.tour)
            col_presence_base = 8
            col_points_base = 9
            col_presence = col_presence_base + (tour - 1) * 2
            col_points = col_points_base + (tour - 1) * 2
            col_nom = 2
            col_prenom = 3
            col_club = 6

            for row in range(2, feuille.max_row + 1):
                if feuille.cell(row=row, column=col_presence).value in ("X", "x"):

                    nom = feuille.cell(row=row, column=col_nom).value
                    prenom = feuille.cell(row=row, column=col_prenom).value
                    club = feuille.cell(row=row, column=col_club).value
                    cle = f"{nom} {prenom} {club}"
                    if cle in new_points:
                        feuille.cell(row=row, column=col_points).value = int(new_points[cle])
            wb.save(temp_file)

        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur modification Excel : {e}")
            return

        # ---------------------------------------------------
        # 4️⃣ renvoyer le fichier modifié vers GitHub
        # ---------------------------------------------------
        try:
            with open(temp_file, "rb") as f:
                encoded = base64.b64encode(f.read()).decode()

            payload = {
                "message": f"Mise à jour catégorie {app.tournoi.category}",
                "content": encoded,
                "sha": file_sha,
                "branch": branch
            }

            r = requests.put(url, headers=headers, json=payload)
            if r.status_code in (200, 201):
                messagebox.showinfo("Succès", "Points sauvegardés sur GitHub.")
            else:
                messagebox.showerror("Erreur", f"Upload GitHub impossible :\n{r.text}")

        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur upload GitHub : {e}")
        # nettoyage
        if os.path.exists(temp_file):
            os.remove(temp_file)

    button_frame = tk.Frame(points_window)
    button_frame.pack(fill=tk.X, pady=10)
    save_button = ttk.Button(button_frame, text="Sauvegarder Dans Excel", command=save_points)
    save_button.pack(side=tk.LEFT, padx=10)
    github_button = ttk.Button(button_frame, text="Sauvegarder Dans Github", command=save_points_in_github)
    github_button.pack(side=tk.LEFT, padx=10)
    close_button = ttk.Button(button_frame, text="Fermer", command=points_window.destroy)
    close_button.pack(side=tk.RIGHT, padx=10)

    def on_double_click(event):
        item_id = tree.identify_row(event.y)
        column = tree.identify_column(event.x)
        if not item_id or column != '#4':
            return
        try:
            x, y, width, height = tree.bbox(item_id, column)
        except Exception:
            return
        if width <= 0 or height <= 0:
            return
        cur_value = tree.set(item_id, 'points')
        entry = ttk.Entry(tree)
        def validate(P):
            if P == "" or P == "-":
                return True
            return P.isdigit()
        vcmd = (tree.register(validate), '%P')
        entry.config(validate='key', validatecommand=vcmd)
        entry.place(x=x, y=y, width=width, height=height)
        entry.insert(0, cur_value)
        entry.focus_set()
        def finish_edit(event=None):
            val = entry.get().strip()
            if val == '' or val == '-':
                val = '0'
            try:
                intval = int(val)
            except Exception:
                intval = 0
            tree.set(item_id, 'points', str(intval))
            entry.destroy()
        def cancel_edit(event=None):
            try:
                entry.destroy()
            except Exception:
                pass
        entry.bind('<Return>', finish_edit)
        entry.bind('<FocusOut>', finish_edit)
        entry.bind('<Escape>', cancel_edit)

    tree.bind('<Double-1>', on_double_click)

    tree.column('nom', width=150, anchor='w')
    tree.column('prenom', width=120, anchor='w')
    tree.column('club', width=100, anchor='w')
    tree.column('points', width=80, anchor='center')

    points_window.transient(app.root)
    points_window.grab_set()
    points_window.focus_set()


def save_points_in_excel(app, new_points):
    try:
        fichier_entree = lire_excel_name()
        wb = load_workbook(fichier_entree)
        feuille = wb[app.tournoi.category]
        tour = int(app.tournoi.tour)

        colonnes_initiale_presence = 8
        colonne_initiale_points = 9

        colonne_presence = colonnes_initiale_presence + (tour - 1) * 2
        colonne_point = colonne_initiale_points + (tour - 1) * 2

        colonne_nom = 2
        colonne_prenom = 3
        colonne_club = 6

        for row in range(2, feuille.max_row + 1):
            if feuille.cell(row=row, column=colonne_presence).value in ("X", "x"):
                nom_val = feuille.cell(row=row, column=colonne_nom).value
                prenom_val = feuille.cell(row=row, column=colonne_prenom).value
                club_val = feuille.cell(row=row, column=colonne_club).value
                cle = f"{nom_val} {prenom_val} {club_val}"
                if cle in new_points:
                    points = int(new_points[cle])
                    points_feuille = feuille.cell(row=row, column=colonne_point).value
                    if points_feuille is None:
                        feuille.cell(row=row, column=colonne_point).value = points
                    elif int(points_feuille) != points:
                        response = messagebox.askyesno("Info", f"Il existe deja des valeurs pour {cle}, continuez-vous ?")
                        if response:
                            feuille.cell(row=row, column=colonne_point).value = points
        wb.save(fichier_entree)
        messagebox.showinfo(title="Info", message="Points joueurs sauvegardes")
        app.results_modified = [True]
    except FileNotFoundError as e:
        print(e)
    except Exception as e:
        print(f"Erreur : {e}")


def gen_tableaux_croises(app):
    try:
        fichier_excel = lire_excel_name()
        nb_clubs = 11
        wb = load_workbook(fichier_excel)

        categories_config = {
            "Benjamins": {"sheet": "Benjamins", "points_col": "B", "points_line": 45, "presence_col": "B", "presence_line": 7},
            "Poussins": {"sheet": "Poussins", "points_col": "P", "points_line": 45, "presence_col": "P", "presence_line": 7},
            "Minimes": {"sheet": "Minimes", "points_col": "I", "points_line": 45, "presence_col": "I", "presence_line": 7},
            "Feminines": {"sheet": "Feminines", "points_col": "B", "points_line": 64, "presence_col": "B", "presence_line": 25},
            "Cadets-Juniors": {"sheet": "Cadets-Juniors", "points_col": "I", "points_line": 64, "presence_col": "I", "presence_line": 25}
        }

        tours_config = {
            "1": {"col_presence": 8, "col_points": 9},
            "2": {"col_presence": 10, "col_points": 11},
            "3": {"col_presence": 12, "col_points": 13},
            "4": {"col_presence": 14, "col_points": 15}
        }

        for category, config in categories_config.items():
            sheet_name = config["sheet"]
            points_col = config["points_col"]
            points_line = config["points_line"]
            presence_col = config["presence_col"]
            presence_line = config["presence_line"]

            if sheet_name not in wb.sheetnames or "TCD" not in wb.sheetnames:
                messagebox.showerror("Erreur", f"Les onglets '{sheet_name}' ou 'TBC' sont introuvables.")
                return

            feuille_categorie = wb[sheet_name]
            feuille_tbc = wb["TCD"]

            for tour, tour_config in tours_config.items():
                col_tour_presence = tour_config["col_presence"]
                col_tour_points = tour_config["col_points"]

                club_points = {}
                club_presence = {}
                for row in range(7, feuille_categorie.max_row + 1):
                    presence = feuille_categorie.cell(row=row, column=col_tour_presence).value
                    club = feuille_categorie.cell(row=row, column=6).value
                    points = feuille_categorie.cell(row=row, column=col_tour_points).value
                    if presence in ["x", "X"] and club:
                        if not isinstance(points, (int, float)):
                            points = 0
                        club_points[club] = club_points.get(club, 0) + points
                        club_presence[club] = club_presence.get(club, 0) + 1

                for row in range(presence_line, presence_line + nb_clubs):
                    club_tbc = feuille_tbc.cell(row=row, column=1).value
                    col_tour = chr(ord(points_col) + int(tour) - 1)
                    if club_tbc in club_points:
                        feuille_tbc[f"{col_tour}{row}"] = club_points[club_tbc]

                for row in range(points_line, points_line + nb_clubs):
                    club_tbc = feuille_tbc.cell(row=row, column=1).value
                    if club_tbc in club_presence:
                        feuille_tbc[f"{presence_col}{row}"] = club_presence[club_tbc]

        # sections pour youngest players unchanged
        wb.save(fichier_excel)
        messagebox.showinfo("Succes", "Les tableaux croises ont ete mis a jour et sauvegardes.")
    except FileNotFoundError:
        messagebox.showerror("Erreur", f"Le fichier Excel '{fichier_excel}' est introuvable.")
    except Exception as e:
        messagebox.showerror("Erreur", f"Une erreur s'est produite : {e}")


def load_points_config(app):
    """Wrapper that returns the points configuration for the given app."""
    return load_points_config_for_app(app)
