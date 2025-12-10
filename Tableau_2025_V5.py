import tkinter as tk
from tkinter import ttk, messagebox
import json
import math
import sys
from utils import *
from json_access import load_all_tableau, save_tableau
import os
import pandas as pd
from openpyxl import load_workbook
import PIL
from SaisieMatch import validate_input_set,handle_mode_change,update_set_visibility
from utils import lire_nb_tops
from fractions import Fraction
import win32print
import win32ui
'''from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
import platform
'''

class TournamentBracket:
    def __init__(self, root, category, round_number):

        self.root = root
        self.category = category
        self.round_number = round_number
        self.tourInitial = 0
        self.nb_tops = 0
        self.nb_poules = 0
        self.fichier_Tableau = ""
        self.results_modified = [False]
        self.previous_mode = "2 sets"
        self.checkbox_vars = {}
        self.ok_window_open = False  # État pour la fenêtre "Tournoi OK"
        self.ko_window_open = False  # État pour la fenêtre "Tournoi KO"
        self.intertops_window = None  # Référence à la fenêtre "InterTops"
        self.intertops_window_open = False  # État pour la fenêtre "InterTops"
        
        self.joueur1_selected = ""
        self.joueur2_selected = ""
        self.indexMatch = -1
        self.levelMatch = -1
        self.matchOk = False
        
        self.root.title(f"Tournoi {category} Tour {round_number}")
        
        # Conteneur pour les boutons
        frame_buttons = ttk.Frame(self.root)
        frame_buttons.pack(fill="both", expand=True, padx=30, pady=10)

        # Bouton pour "Charger le Tournoi OK"
        self.load_button_ok = ttk.Button(frame_buttons, text="Charger le Tournoi OK", command=lambda: self.load_tournament("Tournoi OK"))
        self.load_button_ok.pack(fill="x", pady=5)

        # Bouton pour "Charger le Tournoi KO"
        self.load_button_ko = ttk.Button(frame_buttons, text="Charger le Tournoi KO", command=lambda: self.load_tournament("Tournoi KO"))
        self.load_button_ko.pack(fill="x", pady=5)

        # Bouton pour "InterTops"
        self.inter_tops_button = ttk.Button(frame_buttons, text="InterTops", command=lambda: self.load_tournament("InterTops"))
        self.inter_tops_button.pack(fill="x", pady=5)
        self.nb_tops = lire_nb_tops(self.category, self.round_number)  # Assurez-vous que cette fonction est définie
        # Désactiver le bouton InterTops si nb_tops < 2
        if self.nb_tops < 2:
            self.inter_tops_button.config(state="disabled")
        
        # Bouton pour "Calcul de points"
        self.calcul_points_button = ttk.Button(frame_buttons, text="Calcul de points", command=self.show_points_window)
        self.calcul_points_button.pack(fill="x", pady=5)
        
        #  bouton pour Gestion des règles de points
        self.config_points_button = ttk.Button(frame_buttons, text="Gestion des points", command=self.show_config_points_window)
        self.config_points_button.pack(fill="x", pady=5)
        
        # bouton pour tableaux croisées
        self.gen_tableaux_croises_button = ttk.Button(frame_buttons, text="Génération Tableaux croisés", command=self.gen_tableaux_croises)
        self.gen_tableaux_croises_button.pack(fill="x", pady=5)

        # Ajuster la taille minimale de la fenêtre principale
        self.root.geometry("300x250")  # Largeur x Hauteur
        
        #self.tableauOK = []
        #self.tableauKO = []
        self.tableauOK = {}
        self.tableauKO = {}
        self.dossards = {}
        self.levelMatches = {}
        self.chk_print_vars = {}
        self.lh = 100
        
        self.data = {}
        
        self.read_config_files()
        
        self.debugStatus = lire_status_debug("config.ini")
                

    def extract_players_by_progression(self, tableau, nbPoints, players_points):
        """
        Extrait les joueurs d'un tableau en suivant la progression des rounds.
        Associe directement chaque joueur au nombre de points dans players_points.
        Retourne la nouvelle valeur de nbPoints une fois la boucle terminée.
        """
        level = 0
        fini = False
        firstRoundFound = False

        while not fini:
            string_level = f"1/{level}"
            # Si on a déjà trouvé un round et que le suivant n'existe pas, on arrête.
            if firstRoundFound and string_level not in tableau:
                fini = True
            else:
                if string_level in tableau:
                    firstRoundFound = True
                    for key, joueurs in tableau[string_level].items():
                        # Joueur1
                        j1 = joueurs.get('joueur1', "")
                        if j1 and j1 != "---" and j1 not in players_points:
                            players_points[j1] = nbPoints
                            print(f"joueur1: {j1} points: {nbPoints}, level: {level}")
                            
                        # Joueur2
                        j2 = joueurs.get('joueur2', "")
                        if j2 and j2 != "---" and j2 not in players_points:
                            players_points[j2] = nbPoints
                            print(f"joueur2: {j2} points: {nbPoints}, level: {level}")

                # Progression de level et mise à jour de nbPoints
                if level == 0:
                    level = 1
                    nbPoints -= 2
                else:
                    level *= 2                                      
                    nbPoints -= 1
        
        return (nbPoints + 2)

    def show_points_window(self):
        """Affiche une fenêtre avec les joueurs et leurs points modifiables."""
        # Créer une nouvelle fenêtre
        points_window = tk.Toplevel(self.root)
        points_window.title("Points des joueurs")

        # Ajouter un cadre pour contenir le Treeview et l'ascenseur
        frame = tk.Frame(points_window)
        frame.pack(fill=tk.BOTH, expand=True)

        # Ajouter un Treeview pour afficher les joueurs et leurs points
        columns = ("nom", "prenom", "points")
        tree = ttk.Treeview(frame, columns=columns, show="headings", height=15)
        tree.heading("nom", text="Nom")
        tree.heading("prenom", text="Prénom")
        tree.heading("points", text="Points")

        # Ajouter un ascenseur vertical
        scrollbar_y = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar_y.set)

        # Ajouter un ascenseur horizontal (si nécessaire)
        scrollbar_x = ttk.Scrollbar(frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(xscrollcommand=scrollbar_x.set)

        # Positionner le Treeview et les ascenseurs
        tree.grid(row=0, column=0, sticky="nsew")
        scrollbar_y.grid(row=0, column=1, sticky="ns")
        scrollbar_x.grid(row=1, column=0, sticky="ew")

        # Configurer le redimensionnement
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        # Ajouter les joueurs et leurs points dans le Treeview
        points = self.calculate_points()
        for joueur, pts in sorted(points.items(), key=lambda x: x[1], reverse=True):
            nom, prenom = joueur.split(" ", 1)  # Supposons que le nom et le prénom sont séparés par un espace
            tree.insert("", tk.END, values=(nom, prenom, pts))

        # Fonction pour sauvegarder les points modifiés
        def save_points():
            new_points = {}
            for item in tree.get_children():
                values = tree.item(item, "values")
                joueur = f"{values[0]} {values[1]}"  # Reconstituer le nom complet
                points = int(values[2])  # Les points sont dans la 3e colonne
                new_points[joueur] = points

            # Sauvegarder les points dans un fichier JSON
            with open("players_points.json", "w", encoding="utf-8") as f:
                json.dump(new_points, f, indent=4, ensure_ascii=False)

            # Sauvegarder les points dans Excel
            self.c(new_points)
            points_window.destroy()

        # Ajouter un bouton pour sauvegarder les points
        button_frame = tk.Frame(points_window)
        button_frame.pack(fill=tk.X, pady=10)
        save_button = ttk.Button(button_frame, text="Sauvegarder Dans Excel", command=save_points)
        save_button.pack(side=tk.LEFT, padx=10)

        # Bouton pour fermer la fenêtre
        close_button = ttk.Button(button_frame, text="Fermer", command=points_window.destroy)
        close_button.pack(side=tk.RIGHT, padx=10)

        # Permettre la modification des points
        def on_double_click(event):
            """Permet de modifier les points en double-cliquant sur une cellule."""
            # Identifier l'élément sélectionné et la colonne cliquée
            item_id = tree.focus()
            column = tree.identify_column(event.x)
            if column == "#3":  # Vérifier si la colonne "points" est cliquée
                # Obtenir les coordonnées de la cellule
                x, y, width, height = tree.bbox(item_id, column)
                value = tree.set(item_id, column)

                # Créer un champ de saisie pour modifier les points
                entry = ttk.Entry(tree, width=10)
                entry.place(x=x, y=y, width=width, height=height)
                entry.insert(0, value)
                entry.focus()

                def save_edit(event):
                    new_value = entry.get()
                    if new_value.isdigit():  # Vérifier que la valeur est un entier
                        tree.set(item_id, column=column, value=new_value)
                    entry.destroy()

                entry.bind("<Return>", save_edit)
                entry.bind("<FocusOut>", lambda e: entry.destroy())

        # Lier le double-clic à la modification des points
        tree.bind("<Double-1>", on_double_click)
        
    def quit_match(self, window):
        if (self.results_modified[0]):
            response = messagebox.askyesno(
                    "Info",
                    f"Ders données n'ont pas été sauvées, continuez-vous ?")
            if response:
                self.results_modified = [False]
                window.destroy()
            else:
                window.focus_set()
        else:
            window.destroy()
        
    def save_points_in_excel(self, new_points):
        """
        Exemple de méthode pour récupérer et afficher les points.
        Vous pouvez ensuite les stocker en base ou dans un fichier.
        """
        config_file = "config.ini"
        try:
            fichier_entree = lire_excel_name()
            print(f"Fichier d'entrée : {fichier_entree}")
            
            df = pd.read_excel(fichier_entree, sheet_name=self.category, header=6)
            
            wb = load_workbook(fichier_entree)
            feuille = wb[self.category]
            tour = int(self.round_number)
            champ_presence = f"Pr. T{tour}"
            df_filtre = df[df[champ_presence].str.lower() == "x"]
           
            colonnes_initiale_presence = 8
            colonne_initiale_points =9
            ligne = 8
            
            
            colonne_presence = colonnes_initiale_presence + (tour - 1)*2
            print (f"colonne_presence:{colonne_presence} tour:{tour} colonne_initiale_points:{colonne_initiale_points}")
            colonne_point = colonne_initiale_points + (tour - 1)*2
            print(f"colonne_point:{colonne_point} tour:{tour} colonne_initiale_points:{colonne_initiale_points}")
            
            colonne_nom = 2
            colonne_prenom = 3
            colonne_club = 6
            print(f"new_points:{new_points}")
            for row in range(2, feuille.max_row + 1):  # on part de la ligne 2 car la 1 est l'entête
                if feuille.cell(row=row, column=colonne_presence).value == "X" or  feuille.cell(row=row, column=colonne_presence).value == "x":
                    nom_val = feuille.cell(row=row, column=colonne_nom).value
                    prenom_val = feuille.cell(row=row, column=colonne_prenom).value
                    club_val = feuille.cell(row=row, column=colonne_club).value
                    print(f"nom_val:{nom_val} prenom_val:{prenom_val} club_val:{club_val}")
                    cle = f"{nom_val} {prenom_val} {club_val}"
                
                    champ_points = f"Points T{tour}"

                    if (cle in new_points):
                        
                        points = int(new_points[cle])
                        print(f"points:{points} cle:{cle}")
                        if (points != None):
                            points_feuille = feuille.cell(row=row, column=colonne_point).value
                            print(f"points_feuille:{points_feuille} points:{points}")
                            if (points_feuille != None):
                                points_feuille = int(points_feuille)
                            if (points_feuille == None):
                                print("Ecrit")
                                feuille.cell(row=row, column=colonne_point).value = points
                            elif (points != points_feuille ):
                                print("on est ici")
                                response = messagebox.askyesno(        
                                "Info",
                                f"Il existe déjà des valeurs pour {cle}, continuez-vous ?"
                                )
                                print(f"response:{response}")
                                if (response):
                                    feuille.cell(row=row, column=colonne_point).value = points                        
                            
            wb.save(fichier_entree)
            messagebox.showinfo(title="Info", message="Points joueurs sauvegardés")
            self.results_modified = [True]
            #window.destroy()
        except FileNotFoundError as e:
            print(e)
        except Exception as e:
            print(f"Erreur : {e}")
            
    def autosave_data(self):
        """
        Sauvegarde silencieuse : on enregistre dans le JSON, 
        mais on ne ferme aucune fenêtre et on ne recharge pas le tableau.
        """
        
        print("In autosave_data")
        data = {
            "category": self.category,
            "round": self.round_number,
            "tourInitial": self.tourInitial,
            "nbTops": self.nb_tops,
            "nbPoules": self.nb_poules,
            "Tableau OK": self.tableauOK,
            "Tableau KO": self.tableauKO
        }
        resp = save_tableau(self.fichier_Tableau, data)
        if resp:
            print("Sauvegarde automatique OK.")
            self.results_modified[0] = False
        else:
            print("Erreur lors de la sauvegarde automatique.")
    
    def load_tournament(self, Tableau):
        # Vérifier si la fenêtre correspondante est déjà ouverte
        if Tableau == "Tournoi OK" and self.ok_window_open:
            self.ok_window.focus_set()  # Donner le focus à la fenêtre "Tournoi OK"
            self.current_table = "Tournoi OK"
            return
        elif Tableau == "Tournoi KO" and self.ko_window_open:
            self.ko_window.focus_set()  # Donner le focus à la fenêtre "Tournoi KO"
            self.current_table = "Tournoi KO"
            return
        elif Tableau == "InterTops" and hasattr(self, "intertops_window_open") and self.intertops_window_open:
            self.intertops_window.focus_set()  # Donner le focus à la fenêtre "InterTops"
            self.current_table = "InterTops"
            return
        
        # Mettre à jour la fenêtre sélectionnée
        self.current_table = Tableau
        
        if self.results_modified[0]:
            self.autosave_data()
    
        self.fichier_Tableau = f"Tableau_{self.category}_tour{self.round_number}.json"
        self.data = load_all_tableau(self.fichier_Tableau, True)
        self.tourInitial = int(self.data.get("tourInitial"))
        self.nb_tops = int(self.data.get("nbTops"))
        self.nb_poules = int(self.data.get("nbPoules"))
        
        if self.nb_tops < 2:
            self.inter_tops_button.config(state="disabled")   
        
        new_tableauOK = self.data.get("Tableau OK", [])
        new_tableauKO = self.data.get("Tableau KO", [])
        
        if isinstance(self.tableauOK, dict) and isinstance(new_tableauOK, dict):
            self.tableauOK.clear()
            self.tableauOK.update(new_tableauOK)
        else:
            self.tableauOK = new_tableauOK

        if isinstance(self.tableauKO, dict) and isinstance(new_tableauKO, dict):
            self.tableauKO.clear()
            self.tableauKO.update(new_tableauKO)
        else:
            self.tableauKO = new_tableauKO
        
        self.current_table = Tableau
        
        # Créer une nouvelle fenêtre Toplevel
        new_window = tk.Toplevel(self.root)
        new_window.title(f"{Tableau} - {self.category} Tour {self.round_number}")
        new_window.geometry("1200x800")  # Taille de la fenêtre
        
        # Ajouter un gestionnaire d'événements pour mettre à jour self.current_table
        new_window.bind("<FocusIn>", lambda e: self.update_current_table(Tableau))

        # Ajouter un label pour indiquer le type de tableau
        label = ttk.Label(new_window, text=f"{self.category} - {Tableau}", font=("Arial", 16, "bold"))
        label.pack(pady=10)

        # Conteneur pour le canevas
        canvas_frame = tk.Frame(new_window)
        canvas_frame.pack(fill="both", expand=True)

        # Canvas pour dessiner le tableau avec défilement
        canvas = tk.Canvas(canvas_frame, width=1000, height=600, bg="white")
        scrollbar = tk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
  
      # Conteneur pour les boutons "Refresh" et "Exporter en PDF"
        button_frame = ttk.Frame(new_window)
        button_frame.pack(pady=10)

        # Bouton "Refresh"
        refresh_button = ttk.Button(button_frame, text="Refresh", command=lambda: self.refresh_tableau(Tableau, canvas))
        refresh_button.pack(side="left", padx=5)

        # Bouton "Exporter en PDF"
        if Tableau in ["Tournoi OK", "Tournoi KO"]:
            pdf_button = ttk.Button(button_frame, text="Exporter en PDF", command=lambda: self.print_bracket_to_pdf(canvas))
            pdf_button.pack(side="left", padx=5)
  		
        firstTour=self.tourInitial
        if (self.nb_tops >= 1):
            firstTour= firstTour * 2
            
        if Tableau == "Tournoi OK":
            self.draw_bracket(self.tableauOK, firstTour, self.nb_tops * 4, table_name="OK", canvas=canvas)
            self.ok_window_open = True  # Marquer la fenêtre comme ouverte
            self.ok_window = new_window  # Stocker la référence à la fenêtre
            self.ok_canvas = canvas  # Stocker le canevas pour le PDF

        elif Tableau == "Tournoi KO":
            self.draw_bracket(self.tableauKO, self.tourInitial, 0, table_name="KO", canvas=canvas)
            self.ko_window_open = True  # Marquer la fenêtre comme ouverte
            self.ko_window = new_window  # Stocker la référence à la fenêtre
            self.ko_canvas = canvas  # Stocker le canevas pour le PDF

        elif Tableau == "InterTops":
            self.draw_intertops(self.tableauOK, canvas=canvas)
            self.intertops_window_open = True  # Marquer la fenêtre comme ouverte
            self.intertops_window = new_window  # Stocker la référence à la fenêtre
            self.intertops_canvas = canvas  # Stocker le canevas pour le PDF
            
        def on_close():
            if Tableau == "Tournoi OK":
                self.ok_window_open = False
                self.ok_window = None  # Réinitialiser la référence
                self.ok_canvas = None  # Réinitialiser le canevas
            elif Tableau == "Tournoi KO":
                self.ko_window_open = False
                self.ko_window = None  # Réinitialiser la référence
                self.ko_canvas = None  # Réinitialiser le canevas
            elif Tableau == "InterTops":
                self.intertops_window_open = False
                self.intertops_window = None  # Réinitialiser la référence
                self.intertops_canvas = None  # Réinitialiser le canevas
            new_window.destroy()

        new_window.protocol("WM_DELETE_WINDOW", on_close)
    
    def update_current_table(self, Tableau):
        """Met à jour self.current_table en fonction de la fenêtre active."""
        self.current_table = Tableau
        
    def refresh_tableau(self, Tableau, canvas):
        """Rafraîchit les données et redessine le tableau."""
        self.data = load_all_tableau(self.fichier_Tableau, True)
        new_tableauOK = self.data.get("Tableau OK", [])
        new_tableauKO = self.data.get("Tableau KO", [])
        
        if isinstance(self.tableauOK, dict) and isinstance(new_tableauOK, dict):
            self.tableauOK.clear()
            self.tableauOK.update(new_tableauOK)
        elif isinstance(self.tableauOK, list) and isinstance(new_tableauOK, list):
            self.tableauOK.clear()
            self.tableauOK.extend(new_tableauOK)
        else:
            self.tableauOK = new_tableauOK

        if isinstance(self.tableauKO, dict) and isinstance(new_tableauKO, dict):
            self.tableauKO.clear()
            self.tableauKO.update(new_tableauKO)
        else:
            self.tableauKO = new_tableauKO
            
        # Recharger les états des cases à cocher
        for (level, match_index), var in self.checkbox_vars.items():
            if level in self.tableauOK and str(match_index) in self.tableauOK[level]:
                var.set(self.tableauOK[level][str(match_index)].get("checkbox_state", False))
            if level in self.tableauKO and str(match_index) in self.tableauKO[level]:
                var.set(self.tableauKO[level][str(match_index)].get("checkbox_state", False))

        if Tableau == "Tournoi OK":
            firstTour = self.tourInitial * 2 if self.nb_tops >= 1 else self.tourInitial
            self.draw_bracket(self.tableauOK, firstTour, self.nb_tops * 4, table_name="OK", canvas=canvas)
        elif Tableau == "Tournoi KO":
            self.draw_bracket(self.tableauKO, self.tourInitial, 0, table_name="KO", canvas=canvas)
        elif Tableau == "InterTops":
            self.draw_intertops(self.tableauOK, canvas=canvas)
    
    def set_new_player(self, level, idx, player, nb_tops, dossard, tableau, intertop=False):
            new_level = ""
            try:
                if (not intertop):
                    current_level = int(level.split("/")[1])  # Extrait la partie numérique
                    next_level = current_level // 2
                    new_level = f"1/{next_level}"
                else:
                    new_level = "1/8"
            except (ValueError, IndexError):
                print(f"Erreur : Niveau '{level}' est invalide.")
            # Initialiser le niveau dans tableauOK s'il n'existe pas
            if new_level not in tableau:
                tableau[new_level] = {}
            
            if (not intertop):
                if (int(current_level) == 2*nb_tops):
                    index_table = str(idx)
                else:
                    index_table = str(1 + ((idx - 1) // 2))
                # Initialiser l'entrée pour ce match si elle n'existe pas
                if str(index_table) not in tableau[new_level]:
                    tableau[new_level][index_table] = {}

                # Mettre à jour ou ajouter le joueur dans le niveau suivant
                if (int(current_level) == 2*nb_tops):
                    tableau[new_level][str(index_table)]["joueur2"] = player
                    tableau[new_level][str(index_table)]["dossard_joueur2"] = dossard
                else:
                    if (idx %2 == 1):
                        tableau[new_level][str(index_table)]["joueur1"] = player
                        print(f"tableau[{new_level}][{index_table}][joueur1]={player}")
                        tableau[new_level][str(index_table)]["dossard_joueur1"] = dossard
                    else:
                        tableau[new_level][str(index_table)]["joueur2"] = player
                        print(f"tableau[{new_level}][{index_table}][joueur2]={player}")
                        tableau[new_level][str(index_table)]["dossard_joueur2"] = dossard
            else:
                if str(idx) not in tableau[new_level]:
                    tableau[new_level][str(idx)] = {}
                tableau[new_level][str(idx)]["joueur1"] = player
                tableau[new_level][str(idx)]["dossard_joueur1"] = dossard
                print(f"{player} placé en position {idx}")

    def gestion_vainqueur(self, player1, player2, dossard1, dossard2, level, index_match, tableau, nb_tops, intertop, scores_sets, match_window):
        if (intertop):
            table_intertops = [(2, 7), (6, 3), (4, 5), (8, 1)]
            #Le 1er du top va en position 2
            #Le 2e du top va en position 7
            #Le 3e du top va en position 6
            #etc ...
            
            # Déterminer les classements en fonction du résultat
            
            if str(index_match + 1) not in tableau["intertops"]:
                tableau["intertops"][str(index_match + 1)] = {}
            position1 = index_match*2 + 1
            position2 = position1 + 1
            if scores_sets.count("1") > scores_sets.count("2"):  # Si player1 gagne
                classement1 = f"{position1}er" if position1 == 1 else f"{position1}ème"
                classement2 = f"{position2}ème"
                msg_clst = f"{classement1} Intertop: {player1}\n{classement2} Intertop: {player2}"
            else:  # Si player2 gagne
                classement1 = f"{position2}ème"
                classement2 = f"{position1}er" if position1 == 1 else f"{position1}ème"
                msg_clst = f"{classement1} Intertop: {player2}\n{classement2} Intertop: {player1}"
            
            self.set_new_player(level, table_intertops[index_match][0], player1, nb_tops, dossard1, tableau, intertop)
            self.set_new_player(level, table_intertops[index_match][1], player2, nb_tops, dossard2, tableau, intertop)

            # Mettre à jour les classements dans le tableau
            
            tableau["intertops"][str(index_match + 1)]["classement_top_joueur1"] = classement1
            tableau["intertops"][str(index_match + 1)]["classement_top_joueur2"] = classement2
            print("classement1:", classement1, " classement2:", classement2, 
                  "classement_top_joueur1:", tableau["intertops"][str(index_match + 1)].get("classement_top_joueur1", "N/A"),
                  "classement_top_joueur2:", tableau["intertops"][str(index_match + 1)].get("classement_top_joueur2", "N/A"))

            # Mettre à jour les scores
            tableau["intertops"][str(index_match + 1)]["score"] = scores_sets

            # Afficher un message avec les classements
            self.root.withdraw()  # Masquer la fenêtre principale
            messagebox.showinfo(title="Info", message=msg_clst)
            self.root.deiconify()  # Réafficher la fenêtre principale
        else:
            self.set_new_player(level, index_match, player1, nb_tops, dossard1, tableau, intertop)
            tableau[level][str(index_match)]["score"] = scores_sets
            self.root.withdraw()  # Masquer la fenêtre principale
            messagebox.showinfo(title="Info", message=f"{player1} vainqueur")
            self.matchOk = True
            self.root.deiconify()  # Réafficher la fenêtre principale
        
        # Remettre le focus sur la fenêtre "Détails du Match"
        match_window.focus_set()        
        
    def handle_enter(self, event, entries, player1, player2, level, index_match, tableau, nb_tops, intertop):
        """Gère la validation des scores et détermine le vainqueur."""
        sets_player1 = 0
        sets_player2 = 0
        self.results_modified = [True]
        mode = self.sets_mode_global.get()
        

        # Navigation entre champs
        current_widget = event.widget
        current_index = entries.index(current_widget)

        # Vérifier que la case précédente est remplie, sauf pour la première case
        if current_widget.get() == "":
            current_widget.insert(0, "0")  # Considérer vide comme 0
        elif current_widget.get() == "-":
            current_widget.insert(0, "-0")  # Considérer "-" comme "-0"
            
        if current_index < len(entries):
            # Validation et calcul des scores
            scores_sets =  []
            for entry in entries:
                value = entry.get()
                if value.startswith("-"):
                    sets_player2 += 1
                elif value.isdigit():
                    sets_player1 += 1
                scores_sets.append(value)
                
               
            if (intertop):
                dossard1 = tableau["intertops"][str(index_match+1)].get("dossard_joueur1", "Unknown")
                dossard2 = tableau["intertops"][str(index_match+1)].get("dossard_joueur2", "Unknown")
            else:
                dossard1 = tableau.get(level, {}).get(str(index_match), {}).get("dossard_joueur1", "Unknown")   
                dossard2 = tableau.get(level, {}).get(str(index_match), {}).get("dossard_joueur2", "Unknown")
                
            match_window = event.widget.master  # Référence à la fenêtre "Détails du Match"
            
            if ( (sets_player1 == 2 and mode == "2 sets") or
                (sets_player1 == 3 and mode == "3 sets") ):                        
                self.gestion_vainqueur(player1, player2, dossard1, dossard2, level, index_match, tableau, nb_tops, intertop, scores_sets, match_window)   
                
            elif ( (sets_player2 == 2 and mode == "2 sets") or
                (sets_player2 == 3 and mode == "3 sets") ):                                          
                self.gestion_vainqueur(player2, player1, dossard2, dossard1, level, index_match, tableau, nb_tops, intertop, scores_sets, match_window)
            else:
                next_widget = entries[current_index + 1]  # Passer au champ suivant
                next_widget.focus_set()                    
    
    def update_set_visibility(self, current_mode, set_entries, window=None):
        new_mode = self.sets_mode_global.get()
        confirmation_message = "Les données actuelles n'ont pas été sauvegardées. Voulez-vous continuer et effacer les données ?"
        success = handle_mode_change(
            current_mode=self.sets_mode_global.get(),
            new_mode=new_mode,
            result_entries=[set_entries],
            results_modified=self.results_modified
        )
        if success:
            self.results_modified = [False]
            self.previous_mode = new_mode
        else:
            self.sets_mode_global.set(self.previous_mode)
        if window != None:
            window.focus_set()
        update_set_visibility(self.previous_mode, set_entries)
           
    def save_results(self, window=None):
        """Enregistre les résultats saisis dans un fichier JSON."""
        print("In save_results, joueur1_selected:", self.joueur1_selected, " joueur2_selected:", self.joueur2_selected)
        file_name = f"Tableau_{self.category}_tour{self.round_number}.json"
        
        # Vérifier si le mode est "2 sets" ou "3 sets" et si le vainqueur est connu
        mode = self.sets_mode_global.get()
        print(f"Mode de sets sélectionné : {mode}")
        if mode in ["2 sets", "3 sets"]:
            if not self.matchOk:
                        messagebox.showerror(
                            "Erreur",
                            f"Le vainqueur n'est pas déterminé pour le match {self.joueur1_selected} vs {self.joueur2_selected}. "
                        )
                        return
     
        if mode in ["Forfaitj1", "Forfaitj2"]:
            print("Forfait détecté :", mode)

            winner =""
            loser =""
            d_win =""
            d_lose =""

            # Vainqueur si Forfait joueur 1 → joueur 2 gagne
            if mode == "Forfaitj1":
                winner = self.joueur2_selected
                loser  = self.joueur1_selected
            # Vainqueur si Forfait joueur 2 → joueur 1 gagne
            elif mode == "Forfaitj2":
                winner = self.joueur1_selected
                loser  = self.joueur2_selected

            d_win = self.dossards.get(winner, "")
            d_lose = self.dossards.get(loser, "")
            scores_sets = ["WO"]   # Score unique : forfait

            # Appel normal à gestion_vainqueur
            print(f"match_index:{self.indexMatch} level:{self.levelMatch} winner:{winner} loser:{loser} d_win:{d_win} d_lose:{d_lose} scores_sets:{scores_sets}")
            self.gestion_vainqueur(
                winner,
                loser,
                d_win,
                d_lose,
                self.levelMatch,
                self.indexMatch,
                self.tableauOK,
                self.nb_tops,
                False,          # pas intertop
                scores_sets,
                window
            )
            
            self.levelMatch = ""
            self.indexMatch = -1
     
        # Charger les données existantes si le fichier existe
        data = {
            "category": self.category,
            "round": self.round_number,
            "tourInitial": self.tourInitial,
            "nbTops": self.nb_tops,
            "nbPoules": self.nb_poules,
            "Tableau OK": self.tableauOK,
            "Tableau KO": self.tableauKO
        }
        # Enregistrer dans le fichier
        
        # Sauvegarder les états des cases à cocher
        for (level, match_index), var in self.checkbox_vars.items():
            if level in self.tableauOK and str(match_index) in self.tableauOK[level]:
                self.tableauOK[level][str(match_index)]["checkbox_state"] = var.get()
            if level in self.tableauKO and str(match_index) in self.tableauKO[level]:
                self.tableauKO[level][str(match_index)]["checkbox_state"] = var.get()
        
        resp = save_tableau(self.fichier_Tableau, data)
        if (resp and window != None):
            window.destroy()
            
            # Rafraîchir toutes les fenêtres ouvertes
            if self.ok_window_open:
                self.refresh_tableau("Tournoi OK", self.ok_canvas)
            if self.ko_window_open:
                self.refresh_tableau("Tournoi KO", self.ko_canvas)
            if self.intertops_window_open:
                self.refresh_tableau("InterTops", self.intertops_canvas)
            
            self.results_modified = [False]
            self.load_tournament(self.current_table)
            
        self.matchOk = False
        
    def open_match_window(self, player1, player2, string_level, index_match, tableau, nb_tops, intertop=False):
        # Créer une nouvelle fenêtre
        print(f"In open_match_window: level:{string_level}")
        match_window = tk.Toplevel(self.root)
        match_window.title("Détails du Match")
        
        self.joueur1_selected = player1
        self.joueur2_selected = player2
        self.indexMatch = index_match
        self.levelMatch = string_level
        
        # Déterminer la fenêtre parent
        if self.current_table == "Tournoi OK":
            parent_window = self.ok_window
        elif self.current_table == "Tournoi KO":
            parent_window = self.ko_window
        elif self.current_table == "InterTops":
            parent_window = self.intertops_window

        # Associer la fenêtre "Détails du Match" à la fenêtre parent
        match_window.transient(parent_window)  # Associer à la fenêtre parent
        match_window.grab_set()  # Rendre la fenêtre modale
        
        # Définir self.current_table sur la fenêtre d'origine
        match_window.bind("<FocusIn>", lambda e: self.update_current_table(self.current_table))

        # Ajouter les noms des joueurs
        tk.Label(match_window, text=f"{player1} vs {player2}", font=("Arial", 14)).pack(pady=10)

        # Radio buttons pour 2 sets ou 3 sets
        
        self.sets_mode_global = tk.StringVar(value=self.previous_mode)
        ttk.Radiobutton(match_window, text="2 sets", variable=self.sets_mode_global, value="2 sets", 
            command=lambda: self.update_set_visibility(self.sets_mode_global.get(), set_entries, match_window)).pack()

        ttk.Radiobutton(match_window, text="3 sets", variable=self.sets_mode_global, value="3 sets", 
            command=lambda: self.update_set_visibility(self.sets_mode_global.get(), set_entries, match_window)).pack()
        
        ttk.Radiobutton(match_window, text="Forfait Joueur 1", variable=self.sets_mode_global, value="Forfaitj1", 
            command=lambda: self.update_set_visibility(self.sets_mode_global.get(), set_entries, match_window)).pack()
        
        ttk.Radiobutton(match_window, text="Forfait Joueur 2", variable=self.sets_mode_global, value="Forfaitj2", 
            command=lambda: self.update_set_visibility(self.sets_mode_global.get(), set_entries, match_window)).pack()

        def on_focus_in(event, current_index):
            if current_index < len(set_entries) - 1:
                for idx in range(0, current_index):
                    if (set_entries[idx].get() == ""):
                        set_entries[idx].focus_set()
                        break 
   
        # Champs pour les scores des sets
        set_entries = []
        for i in range(5):
            row = tk.Frame(match_window)
            row.pack(pady=5)
            tk.Label(row, text=f"Set {i + 1}:", font=("Arial", 10)).pack(side="left")

            entry = tk.Entry(row, width=5, validate="key",
                                  validatecommand=(self.root.register(lambda action, value: validate_input_set(action, value, self.results_modified)
                            ), '%d', '%P'))
            entry.pack(side="left")
            entry.bind("<Return>", lambda event, e=set_entries, p1=player1, p2=player2, lvl=string_level, idx=index_match, nbt=nb_tops: self.handle_enter(event, e, p1, p2, lvl, idx, tableau, nbt, intertop))
            entry.bind("<FocusIn>", lambda e, idx=i: on_focus_in(e, idx))
            set_entries.append(entry)

        button_frame = ttk.Frame(match_window)
        button_frame.pack(fill="x", pady=10)

        # Bouton pour fermer la fenêtre
        quit_button = ttk.Button(button_frame, text="Fermer", command=lambda w=match_window: self.quit_match(w))
        quit_button.pack(side="right", padx=5)

        # Bouton pour valider
        validate_button = ttk.Button(button_frame, text="Valider", command=lambda w=match_window: self.save_results(w))
        validate_button.pack(side="left", padx=5)
        
        # Mettre à jour la visibilité initiale des sets
        self.update_set_visibility("2 sets", set_entries)
        
        # Forcer un redimensionnement
        root.update_idletasks()  # Mettre à jour l'interface
        
    def draw_intertops(self, tableau, canvas):
        index_match = 0
        canvas.delete("all")
        
        if 'intertops' not in tableau:
            tableau['intertops'] = {}
            
        # Ajuster la hauteur dynamique du canevas
        canvas_height = max(600, 400 + 8 * 70 * 2)
        start_x = 20
        player_spacing = canvas_height // 8 // 4
        gap = 10  # Espace entre les cadres
        longueur_rectangle = 180
        longueur_fleche = 100
        
        for idx in range (0, 4):
            joueur1 = ""
            joueur2 = ""
            y1_top = idx * 2 * (player_spacing + gap) + 10
            y1_bottom =  y1_top + player_spacing
            mid_y1 = (y1_top + y1_bottom) / 2
            match_nb = str(idx + 1)
            rect_id_1 = canvas.create_rectangle(start_x, y1_top, start_x + longueur_rectangle, y1_bottom, outline="black", width=1)
            
            if (match_nb in tableau['intertops'] and  "joueur1" in tableau['intertops'][match_nb]):
                joueur1 = tableau['intertops'][match_nb]['joueur1']
                dossard1 = tableau['intertops'][match_nb].get('dossard_joueur1', "")
                canvas.create_text(start_x + 10, y1_top + 15, text=joueur1, anchor="w", font=("Arial", 10))
                
                classement1 = tableau['intertops'][match_nb].get("classement_top_joueur1", "")
                canvas.create_text(start_x + 10, y1_top + 30, text=classement1, anchor="w", font=("Arial", 8, "italic"))
                self.dossards[joueur1] = dossard1
            else:
                dossard1 = ""
                joueur1 = ""
            

            y2_top = y1_bottom + 10
            y2_bottom = y2_top + player_spacing
            
            mid_y2 = (y2_top + y2_bottom) // 2
            rect_id_2 = canvas.create_rectangle(start_x, y2_top, start_x + longueur_rectangle, y2_bottom, outline="black", width=1)
            
            if (match_nb in tableau['intertops'] and  "joueur2" in tableau['intertops'][match_nb]):
                joueur2 = tableau['intertops'][match_nb]['joueur2']
                dossard2 = tableau['intertops'][match_nb].get('dossard_joueur2', "")
                canvas.create_text(start_x + 10, y2_top + 15, text=joueur2, anchor="w", font=("Arial", 10))
                
                classement2 = tableau['intertops'][match_nb].get("classement_top_joueur2", "")
                canvas.create_text(start_x + 10, y2_top + 30, text=classement2, anchor="w", font=("Arial", 8, "italic"))
                self.dossards[joueur2] = dossard2
            else:
                dossard2 = ""
                joueur2 = ""
            
            def on_enter(event, dossard_text, joueur_text):
                """Crée et affiche la fenêtre Toplevel quand la souris entre dans le rectangle."""
                if not dossard_text.strip() and not joueur_text.strip():
                    # Si vide => pas d'infobulle
                    return
    
                self._tooltip_window = tk.Toplevel(canvas)
                self._tooltip_window.wm_overrideredirect(True)  # supprime la barre d’entête
                self._tooltip_window.geometry(f"+{event.x_root+10}+{event.y_root+10}")

                doss = f"Dossard: {dossard_text}" if dossard_text else "Dossard inconnu"
                label = tk.Label(self._tooltip_window, text=doss,
                     background="#ffffe0", relief="solid", borderwidth=1)
                label.pack()

            def on_leave(event):
                """Détruit la fenêtre Toplevel quand la souris sort du rectangle."""
                if getattr(self, "_tooltip_window", None) is not None:
                    self._tooltip_window.destroy()
                    self._tooltip_window = None

            # Lier les événements <Enter> et <Leave> au rectangle
            canvas.tag_bind(rect_id_1, "<Enter>", 
                lambda e, doss=dossard1, jr=joueur1: on_enter(e, doss, jr))
            canvas.tag_bind(rect_id_1, "<Leave>", on_leave)

            canvas.tag_bind(rect_id_2, "<Enter>", 
                lambda e, doss=dossard2, jr=joueur2: on_enter(e, doss, jr))
            canvas.tag_bind(rect_id_2, "<Leave>", on_leave)
            
            mid_y = (mid_y1 + mid_y2) // 2
            
            button_x = start_x + longueur_rectangle + 20
            button_y = mid_y
            state = "normal" if joueur1 and joueur2 else "disabled"
            intertop = True
            btn = tk.Button(canvas, text="Détails", state=state, command=lambda p1=joueur1, p2=joueur2, lvl="InterTops", match_index=idx: self.open_match_window(p1, p2, lvl, match_index, tableau, self.nb_tops, intertop))
            canvas.create_window(button_x, button_y, window=btn)
            
            checkbox_x = button_x + 50  # Décalage pour la case
            var = tk.BooleanVar()
            
            if "intertops" in tableau:
                stored_value = tableau["intertops"].get(str(idx+1), {}).get("checkbox_state", False)
                var.set(stored_value)
            else:
                var.set(False)
    
            def checkbox_change_intertops(var, match_index, tableau):
                """Callback pour mettre à jour le 'checkbox_state' d’un match Intertops."""
                # On s’assure que la structure existe :
                self.results_modified[0] = True
                if "intertops" not in tableau:
                    tableau["intertops"] = {}
                if str(match_index) not in tableau["intertops"]:
                    tableau["intertops"][str(match_index)] = {}
                tableau["intertops"][str(match_index)]["checkbox_state"] = var.get()
            
            chk = tk.Checkbutton(
                canvas, variable=var,state=state, 
                command=lambda v=var, m=idx+1: on_checkbox_change_intertops(v, m, tableau)
                )
            canvas.create_window(checkbox_x, button_y, window=chk)

            # Mémoriser cette variable pour la sauvegarde dans save_results
            self.checkbox_vars[("intertops", idx+1)] = var

    def print_bracket_to_pdf(self, canvas):
        import os
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas as pdf_canvas
        from PIL import Image
        
        pdf_directory = lire_pdf_directory()
        os.makedirs(pdf_directory, exist_ok=True)  # Créer le dossier s'il n'existe pas

        #table_type = "OK" if "OK" in self.current_table else "KO"
        table_type = "OK" if canvas == self.ok_canvas else "KO"

        # Définir le nom du fichier de sortie
        pdf_filename = os.path.join(pdf_directory, f"Tableau_{table_type}_{self.category}_Tour{self.round_number}.pdf")
        
        # 1) Déterminer la zone totale du Canvas
        bbox = canvas.bbox("all")

        # ⚠️ Vérifier si `bbox` est `None`
        if bbox is None:
            messagebox.showerror("Impression", "Rien à imprimer ou dimensions invalides.")
            return

        left, top, right, bottom = bbox
        width_canvas = right - left
        height_canvas = bottom - top

        if width_canvas <= 0 or height_canvas <= 0:
            messagebox.showerror("Impression", "Rien à imprimer ou dimensions invalides.")
            return
    
        # 2) Exporter le canvas en PostScript
        ps_filename = os.path.join(pdf_directory, "temp_bracket.ps")
        try:
            canvas.postscript(file=ps_filename, colormode="color", x=left, y=top, width=width_canvas, height=height_canvas)
        except Exception as e:
            messagebox.showerror("Impression", f"Échec de l'export PostScript : {e}")
            return

        # 3) Convertir PostScript -> Image (PNG) avec une meilleure résolution
        png_filename = os.path.join(pdf_directory, "temp_bracket.png")
        try:
            img = Image.open(ps_filename)
            
            # Convertir l'image en mode RGB si nécessaire
            if img.mode != "RGB":
                img = img.convert("RGB")
                
            # Augmenter la résolution pour meilleure lisibilité
            scale_factor = 4  # Multiplier la taille par 2 pour plus de netteté
            new_width = int(img.width * scale_factor)
            new_height = int(img.height * scale_factor)

            #img = img.resize((new_width, new_height), RESAMPLE_METHOD)
            img.resize((new_width, new_height), Image.Resampling.LANCZOS)
            img.save(png_filename, "PNG", dpi=(300, 300))
            
            img.close()
        except Exception as e:
            messagebox.showerror("Impression", f"Échec de la conversion en PNG : {e}")
            return

        # 4) Générer le PDF avec la bonne mise en page
        try:
            #pdf_canvas = canvas.Canvas(pdf_filename, pagesize=A4)
            pdf = pdf_canvas.Canvas(pdf_filename, pagesize=A4)
            pdf_width, pdf_height = A4  # Dimensions d'une page A4

            # Déterminer l'échelle pour que le tableau rentre sur une page
            scale_x = pdf_width / new_width
            scale_y = pdf_height / new_height
            scale = min(scale_x, scale_y)  # Prendre le plus petit facteur pour tout afficher

            # Calculer les marges pour centrer l'image
            final_width = new_width * scale
            final_height = new_height * scale
            margin_x = (pdf_width - final_width) / 2
            margin_y = (pdf_height - final_height) / 2

            # Ajouter l'image au PDF
            pdf.drawImage(png_filename, margin_x, margin_y, width=final_width, height=final_height)
        
            pdf.showPage()
            pdf.save()

            messagebox.showinfo("Impression", f"PDF généré avec succès : {pdf_filename}")

        except Exception as e:
            messagebox.showerror("Impression", f"Échec de la génération du PDF : {e}")
            return

        # Optionnel : Supprimer les fichiers temporaires
        os.remove(ps_filename)
        os.remove(png_filename)
    

    def gen_tableaux_croises(self):
        """Génère les tableaux croisés en calculant la somme des points et des présences par club pour chaque catégorie."""
        try:
            # Ouvrir le fichier Excel
            fichier_excel = lire_excel_name()
            nb_clubs = 11
            print(f"Fichier Excel : {fichier_excel}")
            wb = load_workbook(fichier_excel)

            # Configuration des catégories
            categories_config = {
                "Benjamins": {"sheet": "Benjamins", "points_col": "B", "points_line": 45, "presence_col": "B", "presence_line": 7},
                "Poussins": {"sheet": "Poussins", "points_col": "P", "points_line": 45, "presence_col": "P", "presence_line": 7},
                "Minimes": {"sheet": "Minimes", "points_col": "I", "points_line": 45, "presence_col": "I", "presence_line": 7},
                "Feminines": {"sheet": "Feminines", "points_col": "B", "points_line": 64, "presence_col": "B", "presence_line": 25},
                "Cadets-Juniors": {"sheet": "Cadets-Juniors", "points_col": "I", "points_line": 64, "presence_col": "I", "presence_line": 25}
            }
                
            tours_config = {
                "1": {"col_presence": 8, "col_points": 9},  # Colonnes H et I
                "2": {"col_presence": 10, "col_points": 11}, # Colonnes J et K
                "3": {"col_presence": 12, "col_points": 13},  # Colonnes L et M
                "4": {"col_presence": 14, "col_points": 15}  # Colonnes N et O
            }

            for category, config in categories_config.items():
                print(f"Catégorie : {category}")
                print(f"Configuration : {config}")

                # Charger la configuration de la catégorie

                sheet_name = config["sheet"]
                points_col = config["points_col"]
                points_line = config["points_line"]
                presence_col = config["presence_col"]
                presence_line = config["presence_line"]

                # Vérifier si les onglets nécessaires existent
                if sheet_name not in wb.sheetnames or "TCD" not in wb.sheetnames:
                    messagebox.showerror("Erreur", f"Les onglets '{sheet_name}' ou 'TBC' sont introuvables.")
                    return
                
                # Charger les feuilles
                feuille_categorie = wb[sheet_name]
                feuille_tbc = wb["TCD"]
                
                for tour, tour_config in tours_config.items():
                    col_tour_presence = tour_config["col_presence"]
                    col_tour_points = tour_config["col_points"]
                    print(f"tour: {tour}, col_tour_presence: {col_tour_presence}, col_tour_points: {col_tour_points}")

                    # Calculer la somme des points et des présences par club
                    club_points = {}
                    club_presence = {}
                    for row in range(7, feuille_categorie.max_row + 1):  # Ligne 7 à la fin
                        presence = feuille_categorie.cell(row=row, column=col_tour_presence).value  
                        club = feuille_categorie.cell(row=row, column=6).value  # Colonne F (club)
                        points = feuille_categorie.cell(row=row, column=col_tour_points).value  

                        if presence in ["x", "X"] and club:
                            if not isinstance(points, (int, float)):
                                points = 0  # Si les points ne sont pas numériques, on considère 0
                            club_points[club] = club_points.get(club, 0) + points
                            if club not in club_presence:
                                club_presence[club] = 0
                            club_presence[club] += 1

                    print(f"Points par club : {club_points}")
                    print(f"Présences par club : {club_presence}")

                    # Mettre à jour les points et les présences dans l'onglet "TBC"
                    for row in range(presence_line, presence_line + nb_clubs):  # Ligne 45 à 55
                        club_tbc = feuille_tbc.cell(row=row, column=1).value  # Colonne A (nom du club)
                        col_offset = ord(points_col) - ord("A")
                        print(f"col_offset: {col_offset}")
                        col_tour = chr(ord(points_col) + int(tour) - 1)
                        print (f"row: {row}, club_tbc: {club_tbc}, col_tour: {col_tour}")
                        
                        if club_tbc in club_points:
                            feuille_tbc[f"{col_tour}{row}"] = club_points[club_tbc]  # Colonne des points
                    
                    for row in range(points_line, points_line + nb_clubs):
                        club_tbc = feuille_tbc.cell(row=row, column=1).value  # Colonne A (nom du club)
                        if club_tbc in club_presence:
                            feuille_tbc[f"{presence_col}{row}"] = club_presence[club_tbc]  # Colonne des présences
            youngest_player_f = None
            youngest_date_f = None

            if "Poussins" in wb.sheetnames:
                feuille_poussins = wb["Poussins"]
                youngest_player = None
                youngest_date = None
                
                columnForTour = 8 + (int(self.round_number) - 1) * 2  # Colonne H + (tour-1)*2

                for row in range(7, feuille_poussins.max_row + 1):  # Ligne 7 à la fin
                    presence = feuille_poussins.cell(row=row, column=columnForTour).value  # Colonne H (présence)
                    birth_date = feuille_poussins.cell(row=row, column=5).value  # Colonne E (date de naissance)
                    nom = feuille_poussins.cell(row=row, column=2).value  # Colonne B (nom)
                    prenom = feuille_poussins.cell(row=row, column=3).value  # Colonne C (prénom)
                    club = feuille_poussins.cell(row=row, column=6).value  # Colonne F (club)
                    sexe = feuille_poussins.cell(row=row, column=4).value  # Colonne D (sexe)

                    if presence in ["x", "X"] and birth_date and sexe == "M":
                        if youngest_date is None or birth_date > youngest_date:
                            youngest_date = birth_date
                            youngest_player = {"nom": nom, "prenom": prenom, "club": club, "birth_date": birth_date}
                            
                    if presence in ["x", "X"] and birth_date and sexe == "F":
                        if youngest_date_f is None or birth_date > youngest_date_f:
                            youngest_date_f = birth_date
                            youngest_player_f = {"nom": nom, "prenom": prenom, "club": club, "birth_date": birth_date}

                if youngest_player:
                    feuille_tbc["A96"] = "Plus jeune joueur présent à l’issue du tour " + str(self.round_number)
                    feuille_tbc["E96"] = youngest_player["nom"]
                    feuille_tbc["F96"] = youngest_player["prenom"]
                    feuille_tbc["G96"] = youngest_player["club"]
                    feuille_tbc["H96"] = youngest_player["birth_date"]
                    
            if "Feminines" in wb.sheetnames:
                feuille_poussins = wb["Poussins"]
                columnForTour = 8 + (int(self.round_number) - 1) * 2  # Colonne H + (tour-1)*2

                for row in range(7, feuille_poussins.max_row + 1):  # Ligne 7 à la fin
                    presence = feuille_poussins.cell(row=row, column=columnForTour).value  # Colonne H (présence)
                    birth_date = feuille_poussins.cell(row=row, column=5).value  # Colonne E (date de naissance)
                    nom = feuille_poussins.cell(row=row, column=2).value  # Colonne B (nom)
                    prenom = feuille_poussins.cell(row=row, column=3).value  # Colonne C (prénom)
                    club = feuille_poussins.cell(row=row, column=6).value  # Colonne F (club)
                    sexe = feuille_poussins.cell(row=row, column=4).value  # Colonne D (sexe)
                            
                    if presence in ["x", "X"] and birth_date and sexe == "F":
                        if youngest_date_f is None or birth_date > youngest_date_f:
                            youngest_date_f = birth_date
                            youngest_player_f = {"nom": nom, "prenom": prenom, "club": club, "birth_date": birth_date}
                    
            if youngest_player_f:
                feuille_tbc["A97"] = "Plus jeune joueuse présente à l’issue du tour " + str(self.round_number)
                feuille_tbc["E97"] = youngest_player_f["nom"]
                feuille_tbc["F97"] = youngest_player_f["prenom"]
                feuille_tbc["G97"] = youngest_player_f["club"]
                feuille_tbc["H97"] = youngest_player_f["birth_date"]

            # Sauvegarder le fichier Excel
            wb.save(fichier_excel)
            messagebox.showinfo("Succès", "Les tableaux croisés ont été mis à jour et sauvegardés.")
            
        except FileNotFoundError:
            messagebox.showerror("Erreur", f"Le fichier Excel '{fichier_excel}' est introuvable.")
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur s'est produite : {e}")
        
        
    def show_config_points_window(self):
        win = tk.Toplevel(self.root)
        win.title("Configuration des règles de points")

        frame = ttk.Frame(win, padding=10)
        frame.pack(fill="both", expand=True)

        config = self.load_points_config()

        # Champs KO
        ttk.Label(frame, text="Tableau KO").grid(row=0, column=0, columnspan=2, pady=5)

        ttk.Label(frame, text="Points de base (1er tour KO)").grid(row=1, column=0, sticky="w")
        base_ko_var = tk.IntVar(value=config["KO"]["base_points"])
        ttk.Entry(frame, textvariable=base_ko_var, width=5).grid(row=1, column=1)

        ttk.Label(frame, text="+ points par tour").grid(row=2, column=0, sticky="w")
        inc_ko_var = tk.IntVar(value=config["KO"]["increment_per_round"])
        ttk.Entry(frame, textvariable=inc_ko_var, width=5).grid(row=2, column=1)

        ttk.Label(frame, text="Bonus vainqueur").grid(row=3, column=0, sticky="w")
        win_ko_var = tk.IntVar(value=config["KO"]["winner_bonus"])
        ttk.Entry(frame, textvariable=win_ko_var, width=5).grid(row=3, column=1)

        # Champs OK
        ttk.Label(frame, text="Tableau OK").grid(row=4, column=0, columnspan=2, pady=5)

        ttk.Label(frame, text="Points perdants 1er tour (relatif au vainqueur KO)").grid(row=5, column=0, sticky="w")
        rel_ok_var = tk.IntVar(value=config["OK"]["relative_to_KO_winner"])
        ttk.Entry(frame, textvariable=rel_ok_var, width=5).grid(row=5, column=1)

        ttk.Label(frame, text="+ points par tour").grid(row=6, column=0, sticky="w")
        inc_ok_var = tk.IntVar(value=config["OK"]["increment_per_round"])
        ttk.Entry(frame, textvariable=inc_ok_var, width=5).grid(row=6, column=1)

        ttk.Label(frame, text="Bonus vainqueur").grid(row=7, column=0, sticky="w")
        win_ok_var = tk.IntVar(value=config["OK"]["winner_bonus"])
        ttk.Entry(frame, textvariable=win_ok_var, width=5).grid(row=7, column=1)

        def save_and_close():
            new_config = {
                "KO": {
                    "base_points": base_ko_var.get(),
                    "increment_per_round": inc_ko_var.get(),
                    "winner_bonus": win_ko_var.get(),
                },
                "OK": {
                    "relative_to_KO_winner": rel_ok_var.get(),
                    "increment_per_round": inc_ok_var.get(),
                    "winner_bonus": win_ok_var.get(),
                }
            }
            self.save_points_config(new_config)
            win.destroy()

        ttk.Button(frame, text="Sauvegarder", command=save_and_close).grid(row=8, column=0, columnspan=2, pady=10)

    def calculate_points(self):
        config = self.load_points_config()
        points = {}

        # --- KO ---
        ko_rounds = list(self.tableauKO.keys())
        s=ko_rounds[0]
        print (f"ko_rounds before filtering: {ko_rounds}, s={s}")
        ko_rounds = [r for r in ko_rounds if "/" in r and r.split("/")[1].isdigit() and int(r.split("/")[1]) > 0]
        ko_rounds = sorted(ko_rounds, key=lambda s: Fraction(s))
        max_x = max(int(r.split("/")[1]) for r in ko_rounds)
        
        nb_ko_tours=1

        while(max_x>1):
            max_x = max_x//2
            nb_ko_tours+=1
        
        base_ko = config["KO"]["base_points"]
        inc_ko = config["KO"]["increment_per_round"]
        bonus_ko = config["KO"]["winner_bonus"]

        print (f"len(ko_rounds)={len(ko_rounds)}, base_ko={base_ko}, inc_ko={inc_ko}, bonus_ko={bonus_ko}")
        for i, round_name in enumerate(ko_rounds):
            print("i=", i, " round_name=", round_name)
            round_matches = self.tableauKO[round_name]
            for match in round_matches.values():
                print(f"joueurs: {match.get('joueur1', '---')} vs {match.get('joueur2', '---')}, score: {match.get('score', [])}")
                for joueur_key in ["joueur1", "joueur2"]:
                    joueur = match.get(joueur_key, "---")
                    if joueur == "---":
                        continue
                    if self.is_loser(match, joueur_key):
                        points[joueur] = base_ko + i * inc_ko
                        print(f"joueur: {joueur}, i: {i}, points={points[joueur]}")
                    elif i == nb_ko_tours - 1 and self.is_winner(match, joueur_key):
                        points[joueur] = base_ko + i * inc_ko + bonus_ko
                        vainqueur_ko_points = points[joueur]
                        
                    print(f"nb_ko_tours={nb_ko_tours}, i={i}")

        #vainqueur_ko_points = max(points.values()) if points else 0

        # --- OK ---
        # On filtre d'abord les éléments qui sont bien de la forme '1/x'
        #ok_rounds = [s for s in list(self.tableauOK.keys()) if s != 'intertops']
        #ok_rounds = sorted(ok_rounds, key=lambda s: Fraction(s))
        ok_rounds = list(self.tableauOK.keys())
        ok_rounds = [s for s in ok_rounds if "/" in s and s.split("/")[1].isdigit() and int(s.split("/")[1]) > 0]
        ok_rounds = sorted(ok_rounds, key=lambda s: Fraction(s))
        max_x = max(int(r.split("/")[1]) for r in ok_rounds)
        
        nb_ok_tours=1

        while(max_x>1):
            max_x = max_x//2
            nb_ok_tours+=1
        
        vainqueur_ko_points = base_ko + inc_ko * (nb_ko_tours - 1) + bonus_ko
        rel_ok = config["OK"]["relative_to_KO_winner"]
        inc_ok = config["OK"]["increment_per_round"]
        bonus_ok = config["OK"]["winner_bonus"]

        base_ok = vainqueur_ko_points + rel_ok
        
        print(f"ok_rounds={ok_rounds}, base_ok={base_ok}, inc_ok={inc_ok}, bonus_ok={bonus_ok}")

        for i, round_name in enumerate(ok_rounds):
            if round_name == "intertops":
                continue
            round_matches = self.tableauOK[round_name]
            for match in round_matches.values():
                print(f"joueurs: {match.get('joueur1', '---')} vs {match.get('joueur2', '---')}, score: {match.get('score', [])}")
                for joueur_key in ["joueur1", "joueur2"]:
                    joueur = match.get(joueur_key, "---")
                    print(f"JOUEUR={joueur}")
                    if joueur == "---":
                        continue
                    if self.is_loser(match, joueur_key):
                        points[joueur] = base_ok + i * inc_ok
                        print(f"joueur: {joueur}, i: {i}, points={points[joueur]}")
                    elif i == nb_ok_tours - 1 and self.is_winner(match, joueur_key):
                        points[joueur] = base_ok + i * inc_ok + bonus_ok
                        
                    print(f"nb_ok_tours={nb_ok_tours}, i={i}")
        return points
    
    def is_winner(self, match, joueur_key):
        """
        Détermine si un joueur est le vainqueur du match.
        Un set est gagné par joueur1 si le score est positif.
        Un set est gagné par joueur2 si le score est négatif.
        Le joueur ayant le plus de sets gagnés remporte le match.
        """
        score = match.get("score", [])
        print(f"is_winner: joueur_key={joueur_key}, score={score}")

        if not score:
            return False  # Aucun score, pas de vainqueur

        sets_joueur1 = 0
        sets_joueur2 = 0

        # Parcourir les scores des sets
        for set_score in score:
            try:
                set_score = int(set_score)  # Convertir le score en entier
                if set_score > 0:
                    sets_joueur1 += 1
                elif set_score < 0:
                    sets_joueur2 += 1
            except ValueError:
                # Ignorer les scores invalides (non numériques)
                continue

        # Déterminer le vainqueur en fonction du nombre de sets gagnés
        if joueur_key == "joueur1" and sets_joueur1 > sets_joueur2:
            return True
        if joueur_key == "joueur2" and sets_joueur2 > sets_joueur1:
            return True

        return False
    
    def is_loser(self, match, joueur_key):
        return not self.is_winner(match, joueur_key)
        
    def draw_bracket(self, tableau, firstTour, nb_players_top, table_name, canvas):
        canvas.delete("all")
        
        nb_players = 0
        if self.nb_poules == 2:
            nb_players = 2
        elif self.nb_poules <= 4:
            nb_players = 4       
        elif self.nb_poules <= 8:
            nb_players = 8
        elif self.nb_poules <= 16:
            nb_players = 16
        elif self.nb_poules <= 32:
            nb_players = 32

        # Ajuster la hauteur dynamique du canevas
        canvas_height = max(600, 300 + nb_players * 50 * 2)
        canvas.config(scrollregion=(0, 0, 1200, canvas_height))

        player_spacing = canvas_height // nb_players // 6
        gap = 5
        next_round_positions = []

        header_text = f"Tableau {self.category} {table_name} - Tour {self.round_number}"
        canvas.create_text(500, 50, text=header_text, font=("Arial", 16, "bold"))
        
        def on_click_print():
            self.print_bracket_to_pdf(canvas)
        
        # Créer des cadres pour les joueurs et connecter les lignes                    
        index_match = 0
        for idx in range (0, nb_players*2):
                y1_top = idx * 2 * (player_spacing + gap) + 60
                y1_bottom =  y1_top + player_spacing
                mid_y = (y1_top + y1_bottom) / 2
                index_match = 0
                next_round_positions.append(mid_y)
            
        new_positions = []
        last_match_y = 0
        start_x = 20
        
        currentTour = firstTour
        
        nb_matches = self.tourInitial*2
            
        string_level = f"1/{currentTour}"
        
        if (string_level in tableau):
            matchesTour = tableau[string_level]
            
        while (nb_matches >=1):
            new_positions = []
            index_player = 1    
            longueur_rectangle = 130  # Réduction de la largeur des rectangles
            longueur_fleche = 50  # Réduction de la longueur des flèches
            
            level_found_in_matches = False
            string_level = f"1/{currentTour}"
            
            if currentTour == 1:
                label_text = "Finale"
            elif currentTour == 0:
                label_text = "Vainqueur"
            else:
                label_text = f"1/{currentTour}"
                
            canvas.create_text(
                start_x + (longueur_rectangle / 2),   # Coordonnée X (le centre de la colonne)
                20,                                   # Coordonnée Y fixe (en haut)
                text=label_text,
                anchor="center",
                font=("Arial", 10, "bold")
            )
            
            '''if (string_level in tableau):
                matches_list = tableau[string_level]
                level_found_in_matches = True
            '''
            idx = 1
            j1 = ""
            j2 = ""
            index_player = 1
            for round_mid_y in next_round_positions:
                string_level = f"1/{currentTour}"
                if (string_level in tableau):
                    level_found_in_matches = True
                if (level_found_in_matches):
                    matches_list = tableau[string_level]
                else:
                    matches_list = []
                key= str(idx)

                def display_entree(round_mid_y, dossard_text, text, level_str, match_idx):
                    y_top = (round_mid_y - player_spacing//2)
                    y_bottom = y_top + player_spacing
                    mid_y = (y_top + y_bottom) // 2
                    fillColor = "white"
                    if (text == "---"):
                        fillColor = "grey"
                        text = ""
                    rect_id = canvas.create_rectangle(start_x, y_top, start_x + longueur_rectangle, y_bottom, outline="black", width=1, fill=fillColor)
                    #if (text != ""):
                    text_id = canvas.create_text(start_x + 3, y_top + 8, text=text, anchor="w", font=("Arial", 9))  # Réduction de la taille du texte                      

                        
                    def on_enter(event):
                        """Crée et affiche la fenêtre Toplevel quand la souris entre dans le rectangle."""
                        # On se souvient du tip dans un attribut (ex: self._tooltip_window) 
                        # pour pouvoir le fermer plus tard.
                        if not dossard_text.strip() and not text.strip():
                            # Si vide => on ne fait rien (pas de tooltip)
                            return
                        
                        self._tooltip_window = tk.Toplevel(canvas)
                        self._tooltip_window.wm_overrideredirect(True)  # Supprime la barre d'entête
                        # On positionne près du curseur : event.x_root, event.y_root
                        self._tooltip_window.geometry(f"+{event.x_root+10}+{event.y_root+10}")

                        # On y place un label avec le dossard
                        doss = f"Dossard: {dossard_text}" if dossard_text else "Dossard inconnu"
                        label = tk.Label(self._tooltip_window, text=doss, 
                         background="#ffffe0", relief="solid", borderwidth=1)
                        label.pack()

                    def on_leave(event):
                        """Détruit la fenêtre Toplevel quand la souris sort du rectangle."""
                        if getattr(self, "_tooltip_window", None) is not None:
                            self._tooltip_window.destroy()
                            self._tooltip_window = None
                    #.tag_bind(rect_id, "<Motion>", on_motion)
                    canvas.tag_bind(rect_id, "<Enter>", on_enter)
                    canvas.tag_bind(rect_id, "<Leave>", on_leave)
                    
                    return y_top, y_bottom, mid_y
                
                def on_checkbox_print_change(var, level, match_index, tableau):
                    print("On checkbox print change called")
                
                def on_checkbox_change(var, level, match_index, tableau, window):
                    print("On checkbox change called")
                    self.results_modified[0] = True
                    if level  in tableau and str(match_index) in tableau[level]:
                        cbox_string = f"cbx_match_{match_index}"
                        tableau[level][str(match_index)]["checkbox_state"] = var.get()
                    self.save_results()    
                    
                if (index_player == 1):
                    if key in matches_list and "joueur1" in matches_list[key]:
                        j1 = matches_list[str(idx)]['joueur1']
                        d1 = matches_list[key].get('dossard_joueur1', '')
                        self.dossards[j1] = d1
                        print(f"dossard1 for {j1}: {d1}")
                    else:
                        j1, d1 = '', ''
                    y1_top,y1_bottom,mid_y1 = display_entree(round_mid_y, d1, j1, string_level, idx)
                    if (nb_matches == 2*nb_players_top):
                        new_positions.append(round_mid_y)
                    index_player = index_player + 1
                elif (index_player == 2):
                    if key in matches_list and "joueur2" in matches_list[key]:
                        j2 = matches_list[str(idx)]['joueur2']
                        d2 = matches_list[key].get('dossard_joueur2', '')
                        self.dossards[j2] = d2
                        print(f"dossard2 for {j2}: {d2}")
                    else:
                        j2, d2 = '', ''
                        
                    string_2_players = f"{j1} {j2}"
                    self.levelMatches[string_2_players] = string_level
                    y2_top,y2_bottom,mid_y2 = display_entree(round_mid_y, d2, j2, string_level, idx)       
                    mid_y = (mid_y1 + mid_y2) // 2
                    index_player = 1
                    canvas.create_line(start_x + longueur_rectangle, mid_y1, start_x + longueur_rectangle + longueur_fleche, mid_y, fill="black")
                    canvas.create_line(start_x + longueur_rectangle, mid_y2, start_x + longueur_rectangle + longueur_fleche, mid_y, fill="black")
                    
                    if (nb_matches == 2*nb_players_top):
                        new_positions.append(round_mid_y)
                    else:
                        new_positions.append(mid_y)
                        
                    if (nb_matches == 1):
                        last_match_y = mid_y
                        
                    button_x = start_x + 10
                    button_y = mid_y
                    
                    state = "normal"
                    if (self.debugStatus=="False" and (j1 == "" or j1 == "---" or j2 == "" or j2 == "---") ):
                        state="disabled"  
                    btn = tk.Button(canvas, text="Détails", state = state, command=lambda p1=j1, p2=j2, lvl=string_level, match_index=idx, nbt=nb_players_top: self.open_match_window(p1, p2, lvl, match_index, tableau, nbt))
                    canvas.create_window(button_x, button_y, window=btn)
                    
                    print_button_x = button_x + 100  # Décalage horizontal pour le bouton "Print"
                    print_btn = tk.Button(
                        canvas,
                        text="Print",
                        state=state,  # Synchroniser l'état avec le bouton "Details"
                        command=lambda p1=j1, p2=j2: self.on_print_and_uncheck(p1, p2, table_name, self.category, string_level)
                        )
                    
                    
                    '''print_btn = tk.Button(
                        canvas,
                        text="Preview",
                        state=state,
                        command=lambda p1=j1, p2=j2: self.preview_match_print(p1, p2, table_name, self.category)
                        )
                    '''
                    canvas.create_window(print_button_x, button_y, window=print_btn)
                    
                    
                    
                    checkbox_x = button_x + 50  # Ajustez pour positionner la case
                    var = tk.BooleanVar()

                    # Récupérer la valeur stockée s’il y en a une (checkbox_state) :
                    cbox_string = f"cbx_match_{idx}"
                    if string_level not in tableau or str(idx) not in tableau[string_level] or "checkbox_state" not in tableau[string_level][str(idx)]:
                        var.set(False)
                    else:
                        var.set(tableau[string_level][str(idx)]["checkbox_state"])

                    chk = tk.Checkbutton(canvas, variable=var, command=lambda elt=var, lvl=string_level, match_index=idx: on_checkbox_change(elt, lvl, match_index, tableau, self.root))
                    chk.configure(state=state)
                    canvas.create_window(checkbox_x, button_y, window=chk)

                    # Mémoriser cette variable pour la sauvegarde
                    self.checkbox_vars[(string_level, idx)] = var
                    
                    checkbox_x = checkbox_x + 80  # Décalage pour la case "Print"
                    var_print = tk.BooleanVar()
                    self.chk_print_vars[(string_level, idx)] = var_print
                    chk_print = tk.Checkbutton(canvas, variable=var_print, command=lambda elt=var_print, lvl=string_level, match_index=idx: on_checkbox_print_change(elt, lvl, match_index, tableau))
                    chk_print.configure(state=state)
                    canvas.create_window(checkbox_x, button_y, window=chk_print)
                    
                    j1 = ""
                    j2 = ""
                        
                    idx = idx + 1
            start_x = start_x + longueur_rectangle + longueur_fleche
            currentTour = currentTour // 2
            next_round_positions = new_positions            
            if (nb_matches == 2*nb_players_top):
                nb_players_top = 0
            else:
                nb_matches = nb_matches / 2
                
    def print_match_details(self, player1, player2):
        """Affiche les détails du match dans la console."""
        print(f"Match entre {player1} et {player2}")

    def on_print_and_uncheck(self, player1, player2, table_name, category, level):
        """Imprime les détails des matchs cochés et décoche toutes les cases chk_print."""
        # Récupérer les cases cochées
        checked_matches = [
         (level, match_index)
            for (level, match_index), var in self.chk_print_vars.items()
            if var.get()
        ]

        hprinter, dc = self.init_printer()
        dc.StartPage()
        line = 100
        
        checked_count = len(checked_matches)
        
        # Si une seule case est cochée, imprimer ce match
        if checked_count == 1 or checked_count == 0:
            self.print_match_details_direct(line, dc, player1, player2, table_name, category, level)
            #self.generate_pdf_with_table_pdf(player1, player2, table_name, category, level)

        # Si plusieurs cases sont cochées, afficher tous les matchs
        else:
            matches = []
            for level, match_index in checked_matches:
                match = self.tableauOK.get(level, {}).get(str(match_index), {})
                if not match:
                    match = self.tableauKO.get(level, {}).get(str(match_index), {})

                player1 = match.get("joueur1", "---")
                player2 = match.get("joueur2", "---")
                matches.append((player1, player2, level))

                print(f"Match à imprimer: {player1} vs {player2} Level: {level}")
                line = self.print_match_details_direct(line, dc, player1, player2, table_name, category, level)
                
                if line > 6000:
                    dc.EndPage()
                    dc.StartPage()
                    line=100
                else:
                    line += self.lh * 2  # Espace entre les matchs

        dc.EndPage()
        self.close_printer(dc, hprinter)
        messagebox.showinfo("Impression", "Les matchs ont été imprimés.")
        
        # Décoche toutes les cases chk_print
        for var in self.chk_print_vars.values():
            var.set(False)

    def init_printer(self):
        # --- Impression Windows ---
        printer_name = win32print.GetDefaultPrinter()
        hprinter = win32print.OpenPrinter(printer_name)
        dc = win32ui.CreateDC()
        dc.CreatePrinterDC(printer_name)
        dc.StartDoc(f"impression Feuilles de match")
        return hprinter, dc
        
    def close_printer(self, dc, hprinter):
        dc.EndDoc()
        dc.DeleteDC()
        win32print.ClosePrinter(hprinter)

    def print_match_details_direct(self, line, dc, player1, player2, table_name, category, level):
        """Imprime un match dans un format tableau lisible (comme dans l'image)."""
        import re

        # Récupérer dossards, nom, prénom, club dans les structures tableau
        def extract_infos(player_string):
            # Format attendu : "Nom Prénom Club" OU "Nom Prénom"
            parts = player_string.split()
            nom = parts[0] if len(parts) > 0 else ""
            prenom = parts[1] if len(parts) > 1 else ""
            club = parts[2] if len(parts) > 2 else ""
            return nom, prenom, club

        nom1, prenom1, club1 = extract_infos(player1)
        nom2, prenom2, club2 = extract_infos(player2)
        doss1 =self.dossards.get(player1, "")
        doss2 =self.dossards.get(player2, "")
        if doss1 == "Unknown":
            doss1 = ""
        if doss2 == "Unknown":
            doss2 = ""
    
        string_2_players = f"{player1} {player2}"
        level = self.levelMatches.get(string_2_players)

        x = 100
        y = line

        # Titre tableau
        dc.TextOut(x, y, f"Tableau {table_name} - {level} - {category}                                 TVC - Tour {self.round_number}")
        y += self.lh * 3

        # En-têtes

        header = f"{'Dossard':<4}   {'Nom':<15}   {'Prénom':<15}   {'Club':<7}1    2    3    4    5"
        #header = "Dossard:<5 Nom:<15 Prénom:<15 Club:<6 Set1   Set2   Set3   Set4   Set5"
        
        dc.TextOut(x, y, header)
        y += self.lh
        
        dc.TextOut(x, y, f"______________________________________________________________________________")
        y += self.lh

        # Joueur 1
        line1 = f"|{doss1:<4} | {nom1:<15} | {prenom1:<15} | {club1:<7} | {' ':<2} | {' ':<2} | {' ':<2} | {' ':<2} | {' ':<2} |"
        dc.TextOut(x, y, line1)
        y += self.lh
        
        dc.TextOut(x, y, f"______________________________________________________________________________")
        y += self.lh

        # Joueur 2
        line2 = f"|{doss2:<4} | {nom2:<15} | {prenom2:<15} | {club2:<7} | {' ':<2} | {' ':<2} | {' ':<2} | {' ':<2} | {' ':<2} |"
        dc.TextOut(x, y, line2)
        y += self.lh
        
        dc.TextOut(x, y, f"______________________________________________________________________________")
        y += self.lh

        return y
    
    
    def print_pdf(self, pdf_filename):
        """Imprime un fichier PDF en fonction du système d'exploitation."""
        if platform.system() == "Windows":
            # Commande pour imprimer un PDF sur Windows
            acrobat_path = r"C:\Program Files\Adobe\Acrobat DC\Acrobat\AcroRd32.exe"
            if not os.path.exists(acrobat_path):
                messagebox.showerror("Erreur", "Adobe Acrobat Reader n'est pas installé.")
                return
            try:
                subprocess.run([acrobat_path, "/t", pdf_filename], check=True)
            except Exception as e:
                messagebox.showerror("Erreur", f"Échec de l'impression : {e}")
        elif platform.system() == "Linux":
            # Commande pour imprimer un PDF sur Linux
            try:
                subprocess.run(["lp", pdf_filename], check=True)
            except Exception as e:
                messagebox.showerror("Erreur", f"Échec de l'impression : {e}")
        elif platform.system() == "Darwin":
            # Commande pour imprimer un PDF sur macOS
            try:
                subprocess.run(["lp", pdf_filename], check=True)
            except Exception as e:
                messagebox.showerror("Erreur", f"Échec de l'impression : {e}")
        else:
            messagebox.showerror("Erreur", "Système d'exploitation non pris en charge.")

    def generate_pdf_with_table_pdf(self, player1, player2, table_name, category, level):
        """Génère un PDF avec un tableau contenant les détails du match."""
        pdf_filename = f"Match_{player1}_vs_{player2}.pdf"
        doc = SimpleDocTemplate(pdf_filename, pagesize=A4)

        # Récupérer les informations des joueurs
        def extract_infos(player_string):
            parts = player_string.split()
            nom = parts[0] if len(parts) > 0 else ""
            prenom = parts[1] if len(parts) > 1 else ""
            club = parts[2] if len(parts) > 2 else ""
            return nom, prenom, club

        nom1, prenom1, club1 = extract_infos(player1)
        nom2, prenom2, club2 = extract_infos(player2)
        doss1 = self.dossards.get(player1, "")
        doss2 = self.dossards.get(player2, "")

        # Données pour le tableau
        data = [
            ["Dossard", "Nom", "Prénom", "Club", "Set1", "Set2", "Set3", "Set4", "Set5"],
            [doss1, nom1, prenom1, club1, "", "", "", "", ""],
            [doss2, nom2, prenom2, club2, "", "", "", "", ""],
        ]

        # Créer le tableau
        table = Table(data, colWidths=[50, 100, 100, 100, 50, 50, 50, 50, 50])
        table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 1, colors.black),  # Bordures
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),  # Fond pour les en-têtes
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),  # Alignement centré
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),  # Police des en-têtes
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),  # Police des cellules
        ]))

        # Ajouter le tableau au document
        elements = [table]
        doc.build(elements)

        print(f"PDF généré : {pdf_filename}")
        self.print_pdf(pdf_filename)
    
    def preview_match_print(self, player1, player2, table_name, category):
        """Aperçu à l'écran de ce qui sera imprimé."""

        preview = tk.Toplevel(self.root)
        preview.title("Aperçu avant impression")
        preview.geometry("900x600")

        # Fenêtre scrollable
        frame = tk.Frame(preview)
        frame.pack(fill="both", expand=True)

        canvas = tk.Canvas(frame, bg="white")
        canvas.pack(side="left", fill="both", expand=True)

        scrollbar = tk.Scrollbar(frame, orient="vertical", command=canvas.yview)
        scrollbar.pack(side="right", fill="y")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        string_2_players = f"{player1} {player2}"
        level = self.levelMatches.get(string_2_players)

        # Format approximatif A4 en pixels (pour écran)
        page_width = 800
        page_height = 1130

        x = 50
        y = 60
        lh = 35    # même valeur que pour l'impression → cohérence visuelle

        # --- TITRE ---
        canvas.create_text(x, y,
                       text=f"Tableau {table_name} – {level} – {category}",
                       font=("Arial", 16, "bold"),
                       anchor="nw")
        y += lh * 2

        # EN-TÊTES
        header = "Dossard   Nom        Prénom       Club        Set1   Set2   Set3   Set4   Set5"
        canvas.create_text(x, y, text=header, anchor="nw", font=("Courier New", 12))
        
        header_y = y
        y += lh

        # Extraction nom/prénom/club
        def split_player(p):
            parts = p.split()
            return (parts[0] if len(parts) > 0 else "",
                parts[1] if len(parts) > 1 else "",
                parts[2] if len(parts) > 2 else "")

        nom1, prenom1, club1 = split_player(player1)
        nom2, prenom2, club2 = split_player(player2)

        # On récupère aussi les dossards si déjà connus
        doss1 = doss2 = ""
        lvl = str(level)
        doss1 =self.dossards.get(player1, "")
        doss2 =self.dossards.get(player2, "")

        # Lignes joueurs (Courier pour alignement)
        canvas.create_text(
            x, y,
            text=f"{doss1:<10}{nom1:<12}{prenom1:<12}{club1:<12}",
            font=("Courier New", 12),
            anchor="nw"
        )
        y += lh

        canvas.create_text(
            x, y,
            text=f"{doss2:<10}{nom2:<12}{prenom2:<12}{club2:<12}",
            font=("Courier New", 12),
            anchor="nw"
        )
        
        table_left   = x - 10
        table_top    = header_y - 10
        table_bottom = y + 20
        table_right  = x + 800   # largeur du tableau, ajuste si nécessaire

        canvas.create_rectangle(
            table_left,
            table_top,
            table_right,
            table_bottom,
            outline="black",
            width=2
        )

    def read_config_files(self):
        self.fichier_Tableau = f"Tableau_{self.category}_tour{self.round_number}.json"
        print(f"FICHIER TABLEAU: {self.fichier_Tableau}")
        #self.data = load_all_tableau(self.fichier_Tableau, True)
        
    def restore_focus_to_main_window(self):
        """Remet le focus sur la fenêtre principale (Tableau OK ou Tableau KO)."""
        if self.current_table == "Tournoi OK" and self.ok_window:
            self.ok_window.focus_set()
        elif self.current_table == "Tournoi KO" and self.ko_window:
            self.ko_window.focus_set()
            
    # --- CONFIGURATION DES POINTS ---
    def load_points_config(self):
        config_dir = "Json"
        os.makedirs(config_dir, exist_ok=True)
        config_file = os.path.join(config_dir, f"points_config_{self.category}.json")

        if os.path.exists(config_file):
            with open(config_file, "r", encoding="utf-8") as f:
                return json.load(f)
        else:
            # Valeurs par défaut
            return {
                "KO": {"base_points": 1, "increment_per_round": 1, "winner_bonus": 2},
                "OK": {"relative_to_KO_winner": 0, "increment_per_round": 1, "winner_bonus": 2}
        }
    
    def save_points_config(self, config):
        config_dir = "Json"
        os.makedirs(config_dir, exist_ok=True)
        config_file = os.path.join(config_dir, f"points_config_{self.category}.json")

        with open(config_file, "w", encoding="utf-8") as f:
            json.dump(config, f, indent=4, ensure_ascii=False)
                
if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage : python Saisie-Tournoi-V1.5.py <categorie> <tour>")
        sys.exit(1)

    category_arg = sys.argv[1]
    round_arg = sys.argv[2]    
        
    root = tk.Tk()
    app = TournamentBracket(root, category_arg, round_arg)
    root.mainloop()
